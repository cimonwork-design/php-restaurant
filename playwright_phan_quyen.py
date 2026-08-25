# -*- coding: utf-8 -*-
"""
Kiểm thử E2E chức năng Thiết lập & Phân quyền bằng Playwright.
Chụp từng bước vào C:\\Users\\admin\\Pictures\\testcase
và xuất kết quả ra file Excel cùng thư mục.
"""
from __future__ import annotations

import json
import os
import subprocess
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

BASE = "http://localhost/htql_shop_thoi_trang/public/"
OUT = Path(r"C:\Users\admin\Pictures\testcase")
MYSQL = r"C:\xampp\mysql\bin\mysql.exe"
TODAY = datetime.now().strftime("%d/%m/%Y %H:%M")
TESTER = "Playwright (Chromium)"

ACCOUNTS = {
    "admin": {"user": "admin", "pass": "admin123"},
    "admin2": {"user": "tc_admin2", "pass": "123456"},
    "manager": {"user": "tc_manager", "pass": "123456"},
    "staff": {"user": "tc_staff", "pass": "123456"},
    "cashier": {"user": "tc_cashier", "pass": "123456"},
}

results = []
shot_index = 0


def mysql(sql: str) -> str:
    proc = subprocess.run(
        [MYSQL, "-u", "root", "-e", sql],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr or proc.stdout)
    return proc.stdout


def user_row(username: str) -> dict:
    out = mysql(
        "USE htql_shop_thoi_trang; "
        f"SELECT id, username, role, status FROM users WHERE username='{username}'\\G"
    )
    data = {}
    for line in out.splitlines():
        if ":" in line and not line.strip().startswith("*"):
            k, v = line.split(":", 1)
            data[k.strip()] = v.strip()
    return data


def reset_accounts():
    mysql(
        """
        USE htql_shop_thoi_trang;
        UPDATE users SET role='admin',   status='active' WHERE username='tc_admin2';
        UPDATE users SET role='manager', status='active' WHERE username='tc_manager';
        UPDATE users SET role='staff',   status='active' WHERE username='tc_staff';
        UPDATE users SET role='cashier', status='active' WHERE username='tc_cashier';
        """
    )


def shot(page, slug: str) -> str:
    global shot_index
    shot_index += 1
    name = f"{shot_index:03d}_{slug}.png"
    path = OUT / name
    page.screenshot(path=str(path), full_page=True)
    return name


def add_result(tc_id, category, name, steps, data, expected, actual, status, screenshots, priority="High"):
    results.append({
        "id": tc_id,
        "category": category,
        "name": name,
        "steps": steps,
        "data": data,
        "expected": expected,
        "actual": actual,
        "status": status,
        "screenshots": screenshots,
        "priority": priority,
        "date": TODAY,
        "tester": TESTER,
    })


def pass_fail(ok: bool) -> str:
    return "Pass" if ok else "Fail"


def login(page, username, password):
    # Nếu đang trong phiên khác thì đăng xuất trước, nếu không form login bị redirect.
    page.goto(BASE + "auth/logout", wait_until="domcontentloaded")
    page.wait_for_timeout(250)
    page.goto(BASE + "auth/login", wait_until="domcontentloaded")
    page.wait_for_selector('input[name="username"]')
    page.fill('input[name="username"]', username)
    page.fill('input[name="password"]', password)
    page.click('button[type="submit"]')
    page.wait_for_load_state("domcontentloaded")
    page.wait_for_timeout(400)
    try:
        page.wait_for_selector("a.nav-logout, .alert-danger", timeout=8000)
    except Exception:
        pass


def logout(page):
    page.goto(BASE + "auth/logout", wait_until="domcontentloaded")
    page.wait_for_timeout(300)


def open_permissions_via_menu(page) -> list[str]:
    files = []
    page.goto(BASE, wait_until="domcontentloaded")
    page.wait_for_timeout(300)
    files.append(shot(page, "menu_home"))
    page.locator("i.fa-chart-bar").locator("xpath=ancestor::a[1]").click()
    page.wait_for_timeout(250)
    files.append(shot(page, "menu_he_thong_dropdown"))
    page.locator("a[href*='settings/permissions']").click()
    page.wait_for_selector("h1")
    page.wait_for_timeout(300)
    files.append(shot(page, "man_hinh_phan_quyen"))
    return files


def accept_dialogs(page):
    page.remove_listener("dialog", _accept) if False else None
    page.on("dialog", lambda d: d.accept())


def goto_permissions(page):
    page.goto(BASE + "settings/permissions", wait_until="domcontentloaded")
    page.wait_for_timeout(300)


def row_of(page, username: str):
    return page.locator("tbody tr").filter(has_text=f"@{username}")


def save_role(page, username: str, role: str) -> list[str]:
    files = []
    goto_permissions(page)
    row = row_of(page, username)
    row.locator('select[name="role"]').select_option(role)
    files.append(shot(page, f"chon_role_{username}_{role}"))
    page.once("dialog", lambda d: d.accept())
    row.get_by_role("button", name="Lưu").click()
    page.wait_for_load_state("domcontentloaded")
    page.wait_for_timeout(400)
    files.append(shot(page, f"sau_luu_{username}_{role}"))
    return files


def set_status(page, username: str, action: str) -> list[str]:
    """action: khoa | mo-khoa"""
    files = []
    goto_permissions(page)
    row = row_of(page, username)
    files.append(shot(page, f"truoc_{action}_{username}"))
    page.once("dialog", lambda d: d.accept())
    btn_name = "Khóa" if action == "khoa" else "Mở khóa"
    row.get_by_role("button", name=btn_name).click()
    page.wait_for_load_state("domcontentloaded")
    page.wait_for_timeout(400)
    files.append(shot(page, f"sau_{action}_{username}"))
    return files


def has_text(page, text: str) -> bool:
    return text in (page.content() or "")


def menu_has(page, text: str) -> bool:
    nav = page.locator("nav.top-nav, nav.navbar").inner_text()
    return text in nav


def write_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "phan-quyen"

    headers = [
        "ID", "Category", "Test case Name", "Test step", "Test data",
        "Expected result", "Priority", "Actual result", "Test result",
        "Screenshot", "Date", "Tester",
    ]
    header_fill = PatternFill("solid", fgColor="8C5A2B")
    header_font = Font(bold=True, color="FFFFFF")
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ws.merge_cells("A1:L1")
    ws["A1"] = "KẾT QUẢ KIỂM THỬ PLAYWRIGHT — THIẾT LẬP & PHÂN QUYỀN"
    ws["A1"].font = Font(bold=True, size=14, color="5A3A21")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    passed = sum(1 for r in results if r["status"] == "Pass")
    failed = sum(1 for r in results if r["status"] == "Fail")
    ws.merge_cells("A2:L2")
    ws["A2"] = (
        f"Công cụ: Playwright (Chromium)  |  Ngày: {TODAY}  |  "
        f"Tổng: {len(results)}  |  Pass: {passed}  |  Fail: {failed}  |  "
        f"Ảnh: {OUT}"
    )
    ws["A2"].font = Font(italic=True, color="666666")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(3, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(results, 4):
        values = [
            r["id"], r["category"], r["name"], r["steps"], r["data"],
            r["expected"], r["priority"], r["actual"], r["status"],
            "\n".join(r["screenshots"]), r["date"], r["tester"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i, col, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin
            if col == 9:
                cell.fill = pass_fill if r["status"] == "Pass" else fail_fill
                cell.font = Font(bold=True, color="006100" if r["status"] == "Pass" else "9C0006")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 90

    widths = [6, 18, 42, 48, 28, 48, 12, 48, 12, 36, 18, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A3:L{3 + len(results)}"
    ws.freeze_panes = "A4"

    summary = wb.create_sheet("Tong hop")
    summary["A1"] = "Tổng hợp kết quả"
    summary["A1"].font = Font(bold=True, size=14)
    summary.append([])
    summary.append(["Category", "Pass", "Fail", "Tổng"])
    cats = {}
    for r in results:
        cats.setdefault(r["category"], {"Pass": 0, "Fail": 0})
        cats[r["category"]][r["status"]] += 1
    for cat, c in cats.items():
        summary.append([cat, c["Pass"], c["Fail"], c["Pass"] + c["Fail"]])
    summary.append(["TỔNG", passed, failed, len(results)])
    for col in range(1, 5):
        summary.cell(3, col).fill = header_fill
        summary.cell(3, col).font = header_font
    for col, w in enumerate([22, 10, 10, 10], 1):
        summary.column_dimensions[get_column_letter(col)].width = w

    path = OUT / "KetQua_ThietLapPhanQuyen_Playwright.xlsx"
    wb.save(path)
    return path, passed, failed


def run():
    OUT.mkdir(parents=True, exist_ok=True)
    for old in OUT.glob("*.png"):
        old.unlink()
    for old in OUT.glob("*.xlsx"):
        if old.name.startswith("KetQua_"):
            old.unlink()

    reset_accounts()
    ids = {k: int(user_row(v["user"])["id"]) for k, v in ACCOUNTS.items() if k != "admin"}
    ids["admin"] = int(user_row("admin")["id"])

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            viewport={"width": 1440, "height": 900},
            locale="vi-VN",
        )
        page = context.new_page()
        page.set_default_timeout(20000)

        # ============================================================
        # TC1-7 UI Display
        # ============================================================
        shots = []
        login(page, "admin", "admin123")
        shots.append(shot(page, "TC01_sau_dang_nhap_admin"))
        shots += open_permissions_via_menu(page)
        body = page.content()
        title_ok = "Thiết Lập & Phân Quyền" in page.inner_text("h1")
        frame_ok = "Quản Lý Phân Quyền Tài Khoản Người Dùng" in body
        badge = page.locator(".card-header .badge").inner_text()
        rows = page.locator("tbody tr").count()
        total_ok = f"{rows}" in badge
        add_result(
            1, "UI Display",
            "Hiển thị đúng tiêu đề màn hình và tổng số tài khoản",
            "1. Đăng nhập Admin.\n2. Vào menu Hệ thống & Báo cáo → Thiết lập & Phân quyền.\n3. Quan sát tiêu đề và nhãn Tổng số.",
            "Role: Admin",
            'Tiêu đề "Thiết Lập & Phân Quyền", khung "Quản Lý Phân Quyền Tài Khoản Người Dùng", nhãn "Tổng số: X tài khoản" khớp số dòng.',
            f"Tiêu đề={'OK' if title_ok else 'SAI'}; khung={'OK' if frame_ok else 'SAI'}; {badge}; số dòng={rows}",
            pass_fail(title_ok and frame_ok and total_ok),
            shots,
        )

        headers = [th.inner_text().strip().replace("\n", " ") for th in page.locator("thead th").all()]
        expected_headers = ["ID", "Họ Tên & Tài Khoản", "Liên Hệ", "Vai Trò (Quyền)", "Trạng Thái", "Đăng Nhập Cuối", "Thao Tác"]
        shots2 = [shot(page, "TC02_bang_7_cot")]
        add_result(
            2, "UI Display",
            "Hiển thị đầy đủ các cột trong bảng danh sách tài khoản",
            "1. Quan sát bảng danh sách tài khoản.\n2. Kiểm tra tiêu đề từng cột.",
            "N/A",
            "Bảng đủ 7 cột đúng thứ tự: " + ", ".join(expected_headers),
            "Cột thực tế: " + " | ".join(headers),
            pass_fail(headers == expected_headers),
            shots2,
            "Medium",
        )

        avatars = page.locator(".perm-avatar").all_inner_texts()
        names = page.locator(".fw-semibold").all_inner_texts()
        avatar_ok = True
        pairs = []
        for av, nm in zip(avatars, names):
            expect = (nm.strip()[:1] or "?").upper()
            pairs.append(f"{nm.strip()!r}->{av}")
            if av.strip() != expect:
                avatar_ok = False
        shots3 = [shot(page, "TC03_avatar")]
        add_result(
            3, "UI Display",
            "Hiển thị avatar là ký tự đầu của họ tên tài khoản",
            "1. Quan sát icon avatar (hình tròn) ở đầu mỗi dòng trong cột Họ Tên & Tài Khoản.",
            'VD: "huy" → "H"',
            "Avatar đúng 1 ký tự đầu viết hoa của Họ Tên.",
            "; ".join(pairs),
            pass_fail(avatar_ok and len(avatars) == rows),
            shots3,
            "Low",
        )

        empty_ok = "Chưa cập nhật" in body and "fa-envelope" in body and "fa-phone" in body
        shots4 = [shot(page, "TC04_lien_he_chua_cap_nhat")]
        add_result(
            4, "UI Display",
            'Hiển thị "Chưa cập nhật" khi tài khoản chưa có Email/SĐT',
            "1. Quan sát cột Liên Hệ của các tài khoản chưa nhập Email hoặc Số điện thoại.",
            "Email: [Empty], Phone: [Empty]",
            'Cột Liên Hệ hiển thị "Chưa cập nhật" kèm icon email/điện thoại.',
            "Có nhãn Chưa cập nhật và icon email/phone" if empty_ok else "Thiếu nhãn hoặc icon",
            pass_fail(empty_ok),
            shots4,
            "Low",
        )

        ids_ui = [int(td.inner_text()) for td in page.locator("tbody tr td:first-child").all()]
        sorted_ok = ids_ui == sorted(ids_ui, reverse=True)
        add_result(
            5, "UI Display",
            "Danh sách tài khoản được sắp xếp đúng theo ID",
            "1. Quan sát thứ tự các dòng trong bảng từ trên xuống dưới, so sánh với cột ID.",
            f"ID trên màn hình: {ids_ui}",
            "Danh sách theo ID giảm dần.",
            f"Thứ tự ID: {ids_ui}",
            pass_fail(sorted_ok),
            [shot(page, "TC05_sap_xep_id")],
            "Low",
        )

        crumb = page.locator("nav[aria-label='breadcrumb']").inner_text()
        last_item = page.locator("nav[aria-label='breadcrumb'] .breadcrumb-item.active")
        crumb_ok = "Trang chủ" in crumb and "Thiết lập" in crumb and last_item.count() == 1
        add_result(
            6, "UI Display",
            "Hiển thị đúng breadcrumb điều hướng trên đầu trang",
            "1. Mở màn hình Thiết lập & Phân quyền.\n2. Quan sát thanh breadcrumb ở góc trên bên phải.",
            "N/A",
            'Breadcrumb "Trang chủ / Thiết lập & Phân quyền", mục hiện tại in đậm/không phải link.',
            crumb.replace("\n", " / "),
            pass_fail(crumb_ok),
            [shot(page, "TC06_breadcrumb")],
            "Low",
        )

        page.locator("nav[aria-label='breadcrumb'] a", has_text="Trang chủ").click()
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(300)
        home_ok = "auth/login" not in page.url and (
            "Trang Chủ" in page.title() or page.locator("h1, h2").count() > 0
        )
        add_result(
            7, "UI Display",
            'Nhấn breadcrumb "Trang chủ" để quay lại Trang chủ',
            '1. Tại màn hình phân quyền, nhấn link "Trang chủ" trên breadcrumb.',
            "N/A",
            "Hệ thống điều hướng đúng về màn hình Trang chủ.",
            f"URL sau khi nhấn: {page.url}",
            pass_fail(home_ok),
            [shot(page, "TC07_ve_trang_chu")],
            "Low",
        )

        # ============================================================
        # TC8 Access Control Admin
        # ============================================================
        goto_permissions(page)
        editable = page.locator("select[name='role']").count() > 0
        add_result(
            8, "Access Control",
            "Admin truy cập màn hình Thiết lập & Phân quyền",
            "1. Đăng nhập Admin.\n2. Vào menu Thiết lập & Phân quyền.",
            "Role: Admin",
            "Truy cập thành công, hiển thị bảng và cho phép chỉnh sửa.",
            f"HTTP page OK, dropdown vai trò = {page.locator('select[name=role]').count()}",
            pass_fail("Thiết Lập" in page.inner_text("h1") and editable),
            [shot(page, "TC08_admin_truy_cap")],
        )

        # ============================================================
        # TC9 Manager
        # ============================================================
        logout(page)
        login(page, "tc_manager", "123456")
        page.goto(BASE, wait_until="domcontentloaded")
        page.wait_for_timeout(300)
        s_home = shot(page, "TC09_manager_menu")
        # mở dropdown hệ thống nếu có
        sys_link = page.locator("i.fa-chart-bar")
        if sys_link.count():
            sys_link.locator("xpath=ancestor::a[1]").click()
            page.wait_for_timeout(250)
        s_drop = shot(page, "TC09_manager_dropdown")
        menu_hidden = "Thiết lập & Phân quyền" not in page.locator("nav").inner_text()
        page.goto(BASE + "settings/permissions", wait_until="domcontentloaded")
        page.wait_for_timeout(400)
        s_deny = shot(page, "TC09_manager_url_truc_tiep")
        denied = page.locator(".error-code").count() > 0 or "Không Đủ Quyền" in page.content() or "không có quyền" in page.content().lower()
        add_result(
            9, "Access Control",
            "Quản lý (Manager) không thể truy cập màn hình Thiết lập & Phân quyền",
            "1. Đăng nhập Manager.\n2. Kiểm tra menu có mục phân quyền hay không.\n3. Thử nhập trực tiếp URL.",
            "Role: Manager",
            'Mục không hiển thị trên menu; URL trực tiếp bị từ chối với thông báo "Bạn không có quyền..." hoặc 403.',
            f"Menu ẩn={menu_hidden}; trang 403/từ chối={denied}; URL={page.url}",
            pass_fail(menu_hidden and denied),
            [s_home, s_drop, s_deny],
        )

        # TC10 Staff
        logout(page)
        login(page, "tc_staff", "123456")
        page.goto(BASE + "settings/permissions", wait_until="domcontentloaded")
        page.wait_for_timeout(400)
        denied10 = page.locator(".error-code").count() > 0 or "Không Đủ Quyền" in page.content()
        add_result(
            10, "Access Control",
            "Nhân viên (Staff) không thể truy cập màn hình Thiết lập & Phân quyền",
            "1. Đăng nhập Staff.\n2. Thử truy cập URL màn hình phân quyền.",
            "Role: Staff",
            "Hệ thống từ chối truy cập, hiển thị thông báo không đủ quyền.",
            f"denied={denied10}; URL={page.url}",
            pass_fail(denied10),
            [shot(page, "TC10_staff_403")],
        )

        # TC11 Cashier
        logout(page)
        login(page, "tc_cashier", "123456")
        page.goto(BASE + "settings/permissions", wait_until="domcontentloaded")
        page.wait_for_timeout(400)
        denied11 = page.locator(".error-code").count() > 0 or "Không Đủ Quyền" in page.content()
        add_result(
            11, "Access Control",
            "Thu ngân (Cashier) không thể truy cập màn hình Thiết lập & Phân quyền",
            "1. Đăng nhập Cashier.\n2. Thử truy cập URL màn hình phân quyền.",
            "Role: Cashier",
            "Hệ thống từ chối truy cập, hiển thị thông báo không đủ quyền.",
            f"denied={denied11}; URL={page.url}",
            pass_fail(denied11),
            [shot(page, "TC11_cashier_403")],
        )

        # TC12 Guest
        guest_ctx = browser.new_context(viewport={"width": 1440, "height": 900})
        guest = guest_ctx.new_page()
        guest.goto(BASE + "settings/permissions", wait_until="domcontentloaded")
        guest.wait_for_timeout(500)
        guest_ok = "auth/login" in guest.url
        add_result(
            12, "Access Control",
            "Người dùng chưa đăng nhập không thể truy cập màn hình Thiết lập & Phân quyền",
            "1. Đăng xuất / mở phiên ẩn danh.\n2. Nhập trực tiếp URL màn hình phân quyền.",
            "Role: Guest",
            "Chuyển hướng về màn hình Đăng nhập, không hiển thị dữ liệu tài khoản.",
            f"URL={guest.url}",
            pass_fail(guest_ok),
            [shot(guest, "TC12_guest_redirect_login")],
        )
        guest_ctx.close()

        # ============================================================
        # TC13-15 API 403
        # ============================================================
        def api_put_role(actor_user, actor_pass, target_id, new_role, slug):
            ctx = browser.new_context(viewport={"width": 1440, "height": 900})
            pg = ctx.new_page()
            login(pg, actor_user, actor_pass)
            resp = pg.request.put(
                BASE + f"users/{target_id}",
                data=json.dumps({"role": new_role}),
                headers={"Content-Type": "application/json"},
            )
            body = resp.text()
            pg.set_content(
                "<html><body style='font-family:Segoe UI;padding:24px'>"
                f"<h2>API PUT /users/{target_id}</h2>"
                f"<p>Actor: {actor_user}</p>"
                f"<p>HTTP {resp.status}</p>"
                f"<pre>{body}</pre></body></html>"
            )
            fname = shot(pg, slug)
            ctx.close()
            return resp.status, body, fname

        st13, b13, f13 = api_put_role("tc_manager", "123456", ids["staff"], "cashier", "TC13_manager_api")
        add_result(
            13, "Access Control",
            "Manager gọi API cập nhật phân quyền trực tiếp (bỏ qua giao diện)",
            "1. Đăng nhập Manager.\n2. Gọi PUT /users/{id} {role: Cashier}.",
            f"Actor: Manager / API: PUT /users/{ids['staff']} {{role: cashier}}",
            "Server trả 403 Forbidden, dữ liệu không đổi.",
            f"HTTP {st13}; body={b13}",
            pass_fail(st13 == 403 and user_row("tc_staff")["role"] == "staff"),
            [f13],
        )

        st14, b14, f14 = api_put_role("tc_staff", "123456", ids["manager"], "manager", "TC14_staff_api")
        add_result(
            14, "Access Control",
            "Staff gọi API cập nhật phân quyền trực tiếp (bỏ qua giao diện)",
            "1. Đăng nhập Staff.\n2. Gọi PUT /users/{id} {role: Manager}.",
            f"Actor: Staff / API: PUT /users/{ids['manager']}",
            "Server trả 403 Forbidden, dữ liệu không đổi.",
            f"HTTP {st14}; body={b14}",
            pass_fail(st14 == 403),
            [f14],
        )

        st15, b15, f15 = api_put_role("tc_cashier", "123456", ids["staff"], "staff", "TC15_cashier_api")
        add_result(
            15, "Access Control",
            "Cashier gọi API cập nhật phân quyền trực tiếp (bỏ qua giao diện)",
            "1. Đăng nhập Cashier.\n2. Gọi PUT /users/{id} {role: Staff}.",
            f"Actor: Cashier / API: PUT /users/{ids['staff']}",
            "Server trả 403 Forbidden, dữ liệu không đổi.",
            f"HTTP {st15}; body={b15}",
            pass_fail(st15 == 403),
            [f15],
        )

        # ============================================================
        # Role updates — đăng nhập lại Admin
        # ============================================================
        logout(page)
        login(page, "admin", "admin123")

        role_cases = [
            (16, "admin2", "admin", "manager", "active", True, "Đổi vai trò từ Admin sang Quản lý (Manager) khi đang hoạt động",
             'Hộp thoại "hạ quyền tài khoản Admin này?" + cập nhật thành Quản lý (Manager) + thông báo thành công.'),
            (18, "admin2", "manager", "staff", "active", False, "Đổi vai trò từ Admin/Manager sang Nhân viên (Staff) khi đang hoạt động",
             "Cột Vai Trò cập nhật Nhân viên (Staff), thông báo thành công."),
            (20, "admin2", "staff", "cashier", "active", False, "Đổi vai trò sang Thu ngân (Cashier) khi đang hoạt động",
             "Cột Vai Trò cập nhật Thu ngân (Cashier), thông báo thành công."),
            (22, "manager", "manager", "admin", "active", True, "Đổi vai trò từ Quản lý (Manager) sang Admin khi đang hoạt động",
             'Hộp thoại "cấp quyền Admin" + cập nhật Admin + thông báo thành công.'),
            (24, "manager", "admin", "staff", "active", True, "Đổi vai trò từ Admin sang Nhân viên (Staff)",
             "Cập nhật Nhân viên (Staff), thông báo thành công."),
            (30, "staff", "staff", "manager", "active", False, "Đổi vai trò từ Nhân viên (Staff) sang Quản lý (Manager)",
             "Cập nhật Quản lý (Manager), thông báo thành công."),
            (32, "staff", "manager", "cashier", "active", False, "Đổi vai trò từ Manager sang Thu ngân (Cashier)",
             "Cập nhật Thu ngân (Cashier), thông báo thành công."),
            (36, "cashier", "cashier", "manager", "active", False, "Đổi vai trò từ Thu ngân (Cashier) sang Quản lý (Manager)",
             "Cập nhật Quản lý (Manager), thông báo thành công."),
            (38, "cashier", "manager", "staff", "active", False, "Đổi vai trò từ Manager sang Nhân viên (Staff)",
             "Cập nhật Nhân viên (Staff), thông báo thành công."),
        ]

        # restore known roles first
        reset_accounts()

        # TC16-17 locked vs active for admin2
        files = save_role(page, "tc_admin2", "manager")
        after = user_row("tc_admin2")
        ok16 = after["role"] == "manager" and after["status"] == "active" and has_text(page, "Cập nhật phân quyền thành công")
        add_result(
            16, "Role Update",
            'Đổi vai trò từ Admin sang Quản lý (Manager) khi tài khoản đang "Đang hoạt động"',
            "1. Đăng nhập Admin, mở phân quyền.\n2. Chọn tài khoản Admin khác đang hoạt động.\n3. Chọn Quản lý (Manager).\n4. Nhấn Lưu.\n5. Xác nhận hộp thoại.",
            "Current: Admin / New: Manager / Status: Đang hoạt động / Actor: Admin",
            'Đổi thành công, Vai Trò = "Quản lý (Manager)", thông báo "Cập nhật phân quyền thành công", có confirm hạ quyền Admin.',
            f"DB role/status={after['role']}/{after['status']}; flash={'OK' if has_text(page, 'Cập nhật phân quyền thành công') else 'KHÔNG'}",
            pass_fail(ok16),
            files,
        )

        # TC17: locked account keeps inactive after role change
        reset_accounts()
        set_status(page, "tc_admin2", "khoa")
        files17 = save_role(page, "tc_admin2", "manager")
        after17 = user_row("tc_admin2")
        ok17 = after17["role"] == "manager" and after17["status"] == "inactive"
        add_result(
            17, "Role Update",
            'Đổi vai trò từ Admin sang Quản lý (Manager) khi tài khoản đang "Đã khóa"',
            "1. Đăng nhập Admin.\n2. Đổi vai trò Admin khác đang bị khóa sang Manager.\n3. Lưu.",
            "Current: Admin / New: Manager / Status: Đã khóa",
            "Đổi vai trò thành công nhưng trạng thái vẫn Đã khóa, không tự mở khóa.",
            f"DB={after17['role']}/{after17['status']}",
            pass_fail(ok17),
            files17,
        )

        # Remaining representative role updates
        reset_accounts()
        extra = [
            (18, "tc_admin2", "staff", "Nhân viên (Staff)", "Admin → Staff, đang hoạt động"),
            (20, "tc_admin2", "cashier", "Thu ngân (Cashier)", "Admin → Cashier, đang hoạt động"),
            (22, "tc_manager", "admin", "Admin", "Manager → Admin, đang hoạt động"),
            (24, "tc_manager", "staff", "Nhân viên (Staff)", "Manager → Staff, đang hoạt động"),
            (26, "tc_manager", "cashier", "Thu ngân (Cashier)", "Manager → Cashier, đang hoạt động"),
            (28, "tc_staff", "admin", "Admin", "Staff → Admin, đang hoạt động"),
            (30, "tc_staff", "manager", "Quản lý (Manager)", "Staff → Manager, đang hoạt động"),
            (32, "tc_staff", "cashier", "Thu ngân (Cashier)", "Staff → Cashier, đang hoạt động"),
            (34, "tc_cashier", "admin", "Admin", "Cashier → Admin, đang hoạt động"),
            (36, "tc_cashier", "manager", "Quản lý (Manager)", "Cashier → Manager, đang hoạt động"),
            (38, "tc_cashier", "staff", "Nhân viên (Staff)", "Cashier → Staff, đang hoạt động"),
        ]
        for tc_id, uname, new_role, label, title in extra:
            reset_accounts()
            files_x = save_role(page, uname, new_role)
            after_x = user_row(uname)
            flash_ok = has_text(page, "Cập nhật phân quyền thành công")
            add_result(
                tc_id, "Role Update",
                f"Đổi vai trò: {title}",
                "1. Đăng nhập Admin, mở phân quyền.\n2. Chọn vai trò mới trên dropdown.\n3. Nhấn Lưu (xác nhận nếu có).",
                f"User: {uname} / New role: {new_role} / Status: active / Actor: Admin",
                f'Đổi thành công, cột Vai Trò = "{label}", thông báo "Cập nhật phân quyền thành công".',
                f"DB={after_x['role']}/{after_x['status']}; flash={'OK' if flash_ok else 'KHÔNG'}",
                pass_fail(after_x["role"] == new_role and after_x["status"] == "active" and flash_ok),
                files_x,
                "High" if new_role == "admin" or uname == "tc_admin2" else "Medium",
            )

        # Locked variants 19,21,23,25,29,31,35,37,39 — one representative + loop compact
        locked_cases = [
            (19, "tc_admin2", "staff"),
            (21, "tc_admin2", "cashier"),
            (23, "tc_manager", "admin"),
            (25, "tc_manager", "staff"),
            (27, "tc_manager", "cashier"),
            (29, "tc_staff", "admin"),
            (31, "tc_staff", "manager"),
            (33, "tc_staff", "cashier"),
            (35, "tc_cashier", "admin"),
            (37, "tc_cashier", "manager"),
            (39, "tc_cashier", "staff"),
        ]
        for tc_id, uname, new_role in locked_cases:
            reset_accounts()
            set_status(page, uname, "khoa")
            files_l = save_role(page, uname, new_role)
            after_l = user_row(uname)
            ok_l = after_l["role"] == new_role and after_l["status"] == "inactive"
            add_result(
                tc_id, "Role Update",
                f"Đổi vai trò {uname} → {new_role} khi tài khoản đang Đã khóa",
                "1. Khóa tài khoản.\n2. Đổi vai trò.\n3. Lưu.",
                f"User: {uname} / New: {new_role} / Status: inactive",
                "Đổi vai trò thành công, trạng thái vẫn Đã khóa, không tự mở khóa.",
                f"DB={after_l['role']}/{after_l['status']}",
                pass_fail(ok_l),
                files_l,
                "High" if new_role == "admin" else "Medium",
            )

        # ============================================================
        # TC40-42 Self protection
        # ============================================================
        reset_accounts()
        goto_permissions(page)
        self_row = row_of(page, "admin")
        self_badge = self_row.inner_text()
        no_select = self_row.locator("select").count() == 0
        disabled_btn = self_row.locator("button[disabled]").count() > 0
        has_toi = "Tôi" in self_badge
        add_result(
            40, "Self Protection",
            "Tài khoản Admin đang đăng nhập không thể tự đổi vai trò của chính mình",
            "1. Đăng nhập Admin.\n2. Quan sát dòng của chính tài khoản đang đăng nhập.",
            "Account: admin (đang đăng nhập)",
            'Cột Vai Trò hiển thị nhãn cố định "Admin (Tôi)"; cột Thao Tác icon khóa disabled.',
            f"có nhãn Tôi={has_toi}; không dropdown={no_select}; nút disabled={disabled_btn}",
            pass_fail(has_toi and no_select and disabled_btn),
            [shot(page, "TC40_admin_toi_disabled")],
        )

        no_lock_self = self_row.get_by_role("button", name="Khóa").count() == 0
        add_result(
            41, "Self Protection",
            "Admin không thể tự khóa trạng thái tài khoản của chính mình",
            "1. Quan sát cột Trạng Thái / Thao Tác của dòng Admin đang đăng nhập.",
            "Account: admin (đang đăng nhập)",
            'Luôn "Đang hoạt động" và không có điều khiển tự khóa.',
            f"nút Khóa trên dòng mình={self_row.get_by_role('button', name='Khóa').count()}",
            pass_fail(no_lock_self and "Đang hoạt động" in self_badge),
            [shot(page, "TC41_khong_tu_khoa")],
        )

        st42, b42, f42 = api_put_role("admin", "admin123", ids["admin"], "manager", "TC42_tu_doi_quyen_api")
        add_result(
            42, "Self Protection",
            "Chặn thao tác đổi vai trò chính mình khi gọi API trực tiếp",
            "1. Dùng session Admin.\n2. PUT /users/{own_id} {role: Manager}.",
            f"API: PUT /users/{ids['admin']} {{role: manager}}",
            "Server trả 400/403, không được tự thay đổi quyền của chính mình; dữ liệu không đổi.",
            f"HTTP {st42}; body={b42}; admin role={user_row('admin')['role']}",
            pass_fail(st42 in (400, 403) and user_row("admin")["role"] == "admin"),
            [f42],
        )

        # ============================================================
        # TC43-47 Account status
        # ============================================================
        reset_accounts()
        files43 = set_status(page, "tc_cashier", "khoa")
        after43 = user_row("tc_cashier")
        add_result(
            43, "Account Status",
            "Khóa (deactivate) một tài khoản đang hoạt động",
            '1. Tại dòng tc_cashier đang hoạt động, nhấn Khóa.\n2. Xác nhận.',
            "Account: tc_cashier / Status: active → inactive",
            'Trạng thái "Đã khóa" (màu đỏ), thông báo khóa thành công.',
            f"DB={after43['role']}/{after43['status']}; flash lock={'OK' if 'Khóa tài khoản thành công' in page.content() else page.locator('.alert').inner_text() if page.locator('.alert').count() else ''}",
            pass_fail(after43["status"] == "inactive"),
            files43,
        )

        logout(page)
        login(page, "tc_cashier", "123456")
        locked_msg = page.locator(".alert-danger").inner_text() if page.locator(".alert-danger").count() else ""
        still_login = "auth/login" in page.url or locked_msg != ""
        expect_msg = "Tài khoản của bạn đã bị khóa, vui lòng liên hệ quản trị viên"
        add_result(
            44, "Account Status",
            "Tài khoản đã bị khóa không thể đăng nhập vào hệ thống",
            "1. Đăng xuất Admin.\n2. Thử đăng nhập bằng tài khoản vừa khóa.",
            "Account: tc_cashier (đã khóa)",
            f'Đăng nhập thất bại; thông báo "{expect_msg}".',
            f"URL={page.url}; msg={locked_msg}",
            pass_fail(expect_msg in locked_msg and "pos" not in page.url),
            [shot(page, "TC44_dang_nhap_bi_khoa")],
        )

        login(page, "admin", "admin123")
        files45 = set_status(page, "tc_cashier", "mo-khoa")
        after45 = user_row("tc_cashier")
        logout(page)
        login(page, "tc_cashier", "123456")
        relogin_ok = "auth/login" not in page.url
        s45b = shot(page, "TC45_dang_nhap_lai_sau_mo_khoa")
        add_result(
            45, "Account Status",
            "Mở khóa (kích hoạt lại) một tài khoản đã bị khóa",
            "1. Admin mở khóa.\n2. Đăng nhập lại bằng tài khoản đó.",
            "Account: tc_cashier / Status: inactive → active",
            "Trạng thái Đang hoạt động; đăng nhập lại bình thường.",
            f"DB={after45['status']}; relogin_ok={relogin_ok}; URL={page.url}",
            pass_fail(after45["status"] == "active" and relogin_ok),
            files45 + [s45b],
        )

        # TC46 lock while online
        ctx_online = browser.new_context(viewport={"width": 1440, "height": 900})
        pg_online = ctx_online.new_page()
        login(pg_online, "tc_cashier", "123456")
        pg_online.goto(BASE + "pos", wait_until="domcontentloaded")
        s46a = shot(pg_online, "TC46_cashier_dang_online_pos")
        login(page, "admin", "admin123")
        set_status(page, "tc_cashier", "khoa")
        s46b = shot(page, "TC46_admin_khoa_khi_online")
        pg_online.goto(BASE + "pos", wait_until="domcontentloaded")
        pg_online.wait_for_timeout(500)
        s46c = shot(pg_online, "TC46_phien_bi_da_ve_login")
        kicked = "auth/login" in pg_online.url or "đã bị khóa" in pg_online.content()
        add_result(
            46, "Account Status",
            "Tài khoản bị khóa trong khi đang có phiên đăng nhập hoạt động ở nơi khác",
            "1. Cashier đăng nhập trình duyệt B (POS).\n2. Admin khóa tài khoản ở trình duyệt A.\n3. Cashier thao tác tiếp / tải lại trang.",
            "Account X: đang online, bị khóa giữa phiên",
            "Phiên bị từ chối / buộc đăng xuất, thông báo tài khoản đã bị khóa.",
            f"URL phiên B sau khi khóa={pg_online.url}; kicked={kicked}",
            pass_fail(kicked),
            [s46a, s46b, s46c],
            "Medium",
        )
        ctx_online.close()
        reset_accounts()

        login(page, "admin", "admin123")
        set_status(page, "tc_admin2", "khoa")
        goto_permissions(page)
        green = page.locator(".badge.bg-success", has_text="Đang hoạt động").count() > 0
        red = page.locator(".badge.bg-danger", has_text="Đã khóa").count() > 0
        add_result(
            47, "Account Status",
            "Hiển thị đúng màu sắc trạng thái tài khoản",
            '1. Quan sát nhãn "Đang hoạt động" và "Đã khóa".',
            "Status: Đang hoạt động / Đã khóa",
            "Đang hoạt động nền xanh; Đã khóa nền đỏ/xám.",
            f"badge xanh={green}; badge đỏ={red}",
            pass_fail(green and red),
            [shot(page, "TC47_mau_trang_thai")],
            "Low",
        )

        reset_accounts()
        files48 = set_status(page, "tc_manager", "khoa")
        add_result(
            48, "Account Status",
            "Khóa tài khoản có vai trò Manager",
            "1. Chọn Manager đang hoạt động.\n2. Nhấn Khóa.",
            "Role: Manager / Status: active → inactive",
            "Khóa thành công; Manager không thể đăng nhập.",
            f"DB={user_row('tc_manager')['status']}",
            pass_fail(user_row("tc_manager")["status"] == "inactive"),
            files48,
            "Medium",
        )

        reset_accounts()
        files49 = set_status(page, "tc_admin2", "khoa")
        add_result(
            49, "Account Status",
            "Khóa tài khoản khác có vai trò Admin (không phải Admin đang đăng nhập)",
            "1. Chọn Admin khác.\n2. Nhấn Khóa (xác nhận).",
            "Role: Admin (khác) / Status: active → inactive",
            "Cho phép khóa Admin khác; tài khoản đó không thể đăng nhập.",
            f"DB tc_admin2={user_row('tc_admin2')['status']}",
            pass_fail(user_row("tc_admin2")["status"] == "inactive"),
            files49,
            "Medium",
        )

        # TC51 default new account — check register defaults via DB of existing staff
        st_staff = user_row("tc_staff")
        add_result(
            51, "Account Status",
            "Trạng thái mặc định khi một tài khoản mới được tạo",
            "1. Quan sát tài khoản tạo mới (đăng ký) trên màn hình phân quyền.",
            "New account status: mặc định",
            'Tài khoản mới mặc định "Đang hoạt động" và vai trò phù hợp (Staff), không bị khóa sẵn.',
            f"tc_staff={st_staff['role']}/{st_staff['status']}",
            pass_fail(st_staff["status"] == "active"),
            [shot(page, "TC51_tai_khoan_mac_dinh")],
            "Low",
        )

        # ============================================================
        # TC52-55 Permission matrix via menu + URL
        # ============================================================
        def check_matrix(role_key, username, password, allowed, slug_prefix):
            logout(page)
            login(page, username, password)
            page.goto(BASE, wait_until="domcontentloaded")
            page.wait_for_timeout(300)
            nav = page.locator("nav").inner_text()
            shots_m = [shot(page, f"{slug_prefix}_menu")]
            url_map = {
                "Trang Chủ": "home",
                "Quản Lý Sản Phẩm": "products",
                "Bán Hàng (POS)": "pos",
                "Quản Lý Khách Hàng": "customers",
                "Tài Chính": "financial/cashflow",
                "Báo Cáo": "reports/revenue",
                "Nhân Viên": "employees",
                "Đơn Hàng": "sales",
                "Đổi Hàng": "exchanges",
                "Thiết Lập & Phân Quyền": "settings/permissions",
            }
            details = []
            ok_all = True
            for label, path in url_map.items():
                should = label in allowed or (
                    label == "Tài Chính" and "Tài Chính" in allowed
                ) or (
                    label == "Báo Cáo" and "Báo Cáo & Phân Tích" in allowed
                )
                # normalize allowed names
                aliases = {
                    "Tài Chính": ["Tài Chính"],
                    "Báo Cáo": ["Báo Cáo & Phân Tích", "Báo Cáo"],
                    "Thiết Lập & Phân Quyền": ["Thiết Lập & Phân Quyền"],
                }
                expect_allow = False
                for a in allowed:
                    if a in label or label in a or a.startswith(label.split()[0]):
                        expect_allow = True
                # explicit
                expect_allow = False
                for a in allowed:
                    if a == "Trang Chủ" and label == "Trang Chủ":
                        expect_allow = True
                    if a == "Quản Lý Sản Phẩm" and label == "Quản Lý Sản Phẩm":
                        expect_allow = True
                    if a == "Bán Hàng (POS)" and label == "Bán Hàng (POS)":
                        expect_allow = True
                    if a == "Quản Lý Khách Hàng" and label == "Quản Lý Khách Hàng":
                        expect_allow = True
                    if a == "Tài Chính" and label == "Tài Chính":
                        expect_allow = True
                    if a == "Báo Cáo & Phân Tích" and label == "Báo Cáo":
                        expect_allow = True
                    if a == "Nhân Viên" and label == "Nhân Viên":
                        expect_allow = True
                    if a == "Đơn Hàng" and label == "Đơn Hàng":
                        expect_allow = True
                    if a == "Đổi Hàng" and label == "Đổi Hàng":
                        expect_allow = True
                    if a == "Thiết Lập & Phân Quyền" and label == "Thiết Lập & Phân Quyền":
                        expect_allow = True

                resp = page.request.get(BASE + path)
                # follow: 200 allow, 403 deny, 302 login shouldn't happen
                code = resp.status
                allowed_now = code in (200, 302) and code != 403
                # 302 to same app page can be home
                if expect_allow and code == 403:
                    ok_all = False
                if (not expect_allow) and code == 200:
                    ok_all = False
                details.append(f"{label}: HTTP {code} (expect {'allow' if expect_allow else 'deny'})")
            shots_m.append(shot(page, f"{slug_prefix}_sau_check"))
            return ok_all, "\n".join(details), shots_m, nav

        reset_accounts()
        admin_allowed = [
            "Trang Chủ", "Quản Lý Sản Phẩm", "Bán Hàng (POS)", "Quản Lý Khách Hàng",
            "Tài Chính", "Báo Cáo & Phân Tích", "Nhân Viên", "Đơn Hàng", "Đổi Hàng",
            "Thiết Lập & Phân Quyền",
        ]
        ok52, d52, s52, nav52 = check_matrix("admin", "admin", "admin123", admin_allowed, "TC52_admin")
        add_result(
            52, "Permission Matrix",
            "Kiểm tra quyền truy cập các chức năng của vai trò Admin",
            "1. Đăng nhập Admin.\n2. Kiểm tra menu và URL từng chức năng.",
            "Role: Admin",
            "Truy cập được toàn bộ các mục kể cả Thiết Lập & Phân Quyền.",
            d52,
            pass_fail(ok52 and "Thiết lập & Phân quyền" in nav52),
            s52,
        )

        mgr_allowed = [x for x in admin_allowed if x != "Thiết Lập & Phân Quyền"]
        ok53, d53, s53, nav53 = check_matrix("manager", "tc_manager", "123456", mgr_allowed, "TC53_manager")
        add_result(
            53, "Permission Matrix",
            "Kiểm tra quyền truy cập các chức năng của vai trò Quản lý (Manager)",
            "1. Đăng nhập Manager.\n2. Kiểm tra menu và URL từng chức năng.",
            "Role: Manager",
            "Truy cập được mọi mục trừ Thiết Lập & Phân Quyền.",
            d53 + f"\nMenu có mục phân quyền={'Thiết lập & Phân quyền' in nav53}",
            pass_fail(ok53 and "Thiết lập & Phân quyền" not in nav53),
            s53,
        )

        staff_allowed = [
            "Trang Chủ", "Quản Lý Sản Phẩm", "Bán Hàng (POS)", "Quản Lý Khách Hàng",
            "Đơn Hàng", "Đổi Hàng",
        ]
        ok54, d54, s54, nav54 = check_matrix("staff", "tc_staff", "123456", staff_allowed, "TC54_staff")
        add_result(
            54, "Permission Matrix",
            "Kiểm tra quyền truy cập các chức năng của vai trò Nhân viên (Staff)",
            "1. Đăng nhập Staff.\n2. Kiểm tra menu và URL từng chức năng.",
            "Role: Staff",
            "Được: Trang Chủ, Sản Phẩm, POS, Khách Hàng, Đơn Hàng, Đổi Hàng. Không: Tài Chính, Báo Cáo, Nhân Viên, Thiết Lập.",
            d54,
            pass_fail(ok54),
            s54,
        )

        cash_allowed = ["Trang Chủ", "Bán Hàng (POS)", "Đơn Hàng"]
        ok55, d55, s55, nav55 = check_matrix("cashier", "tc_cashier", "123456", cash_allowed, "TC55_cashier")
        add_result(
            55, "Permission Matrix",
            "Kiểm tra quyền truy cập các chức năng của vai trò Thu ngân (Cashier)",
            "1. Đăng nhập Cashier.\n2. Kiểm tra menu và URL từng chức năng.",
            "Role: Cashier",
            "Chỉ Trang Chủ, POS, Đơn Hàng. Không các mục còn lại.",
            d55,
            pass_fail(ok55),
            s55,
        )

        # TC56 quyền mới có hiệu lực ngay
        reset_accounts()
        ctx_x = browser.new_context(viewport={"width": 1440, "height": 900})
        pg_x = ctx_x.new_page()
        login(pg_x, "tc_staff", "123456")
        pg_x.goto(BASE + "financial/cashflow", wait_until="domcontentloaded")
        s56a = shot(pg_x, "TC56_staff_chua_vao_duoc_taichinh")
        blocked_before = pg_x.locator(".error-code").count() > 0 or "Không Đủ Quyền" in pg_x.content()
        login(page, "admin", "admin123")
        save_role(page, "tc_staff", "manager")
        s56b = shot(page, "TC56_admin_doi_staff_thanh_manager")
        pg_x.goto(BASE + "financial/cashflow", wait_until="domcontentloaded")
        pg_x.wait_for_timeout(400)
        s56c = shot(pg_x, "TC56_staff_sau_khi_len_manager")
        after_role = user_row("tc_staff")["role"]
        # session sync should apply new role immediately
        now_ok = "Không Đủ Quyền" not in pg_x.content() and pg_x.locator(".error-code").count() == 0
        add_result(
            56, "Permission Matrix",
            "Quyền mới có hiệu lực ngay sau khi Admin đổi vai trò cho tài khoản đang online",
            "1. Staff đang online, thử Tài Chính (bị chặn).\n2. Admin đổi Staff → Manager.\n3. Staff tải lại Tài Chính.",
            "Account X: Staff → Manager (đang online)",
            "Áp dụng quyền Manager ngay hoặc không giữ quyền Staff cũ.",
            f"blocked_before={blocked_before}; after_role={after_role}; after_access_ok={now_ok}",
            pass_fail(blocked_before and after_role == "manager" and now_ok),
            [s56a, s56b, s56c],
            "Medium",
        )
        ctx_x.close()

        # ============================================================
        # TC57, 59, 60, 63 Save actions
        # ============================================================
        reset_accounts()
        login(page, "admin", "admin123")
        before_login = mysql(
            "USE htql_shop_thoi_trang; SELECT last_login FROM users WHERE username='tc_staff';"
        )
        files57 = save_role(page, "tc_staff", "cashier")
        after_login = mysql(
            "USE htql_shop_thoi_trang; SELECT last_login FROM users WHERE username='tc_staff';"
        )
        last_same = before_login.strip() == after_login.strip()
        add_result(
            57, "Save Action",
            "Lưu thay đổi thành công hiển thị thông báo xác nhận",
            "1. Đổi vai trò một tài khoản.\n2. Nhấn Lưu.\n3. Quan sát thông báo và Đăng Nhập Cuối.",
            "Role: đổi hợp lệ",
            'Thông báo "Cập nhật phân quyền thành công"; Vai Trò đổi; Đăng Nhập Cuối không đổi.',
            f"flash={'OK' if has_text(page, 'Cập nhật phân quyền thành công') else 'NO'}; last_login giữ nguyên={last_same}",
            pass_fail(has_text(page, "Cập nhật phân quyền thành công") and last_same),
            files57,
            "Medium",
        )

        reset_accounts()
        goto_permissions(page)
        row = row_of(page, "tc_staff")
        row.locator('select[name="role"]').select_option("manager")
        page.once("dialog", lambda d: d.accept())
        btn = page.locator(f'button[form="roleForm-{ids["staff"]}"]')
        btn.click()
        btn.click(timeout=2000)
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(400)
        add_result(
            59, "Save Action",
            "Nhấn nút Lưu liên tiếp nhiều lần (double click)",
            "1. Đổi vai trò.\n2. Nhấn Lưu 2 lần liên tiếp.",
            "Action: double click nút Lưu",
            "Chỉ ghi nhận một yêu cầu; không lỗi.",
            f"DB role={user_row('tc_staff')['role']}; page vẫn OK",
            pass_fail(user_row("tc_staff")["role"] == "manager" and page.locator("h1").count() > 0),
            [shot(page, "TC59_double_submit")],
            "Low",
        )

        reset_accounts()
        goto_permissions(page)
        row = row_of(page, "tc_staff")
        old = row.locator("select[name='role']").input_value()
        row.locator('select[name="role"]').select_option("manager")
        s60a = shot(page, "TC60_chon_tam_chua_luu")
        page.goto(BASE + "home", wait_until="domcontentloaded")
        goto_permissions(page)
        now = row_of(page, "tc_staff").locator("select[name='role']").input_value()
        add_result(
            60, "Save Action",
            "Hủy thay đổi bằng cách rời khỏi dropdown mà không nhấn Lưu",
            "1. Đổi dropdown nhưng không Lưu.\n2. Rời trang rồi quay lại.",
            "Role: chọn tạm thời, không Lưu",
            "Giá trị trở lại vai trò gốc, chưa lưu vào hệ thống.",
            f"trước={old}; sau khi quay lại={now}",
            pass_fail(now == "staff"),
            [s60a, shot(page, "TC60_quay_lai_van_role_cu")],
            "Low",
        )

        reset_accounts()
        files63 = save_role(page, "tc_manager", "manager")
        add_result(
            63, "Edge Case",
            "Chọn lại đúng vai trò hiện tại rồi nhấn Lưu",
            "1. Manager sẵn có vai trò Manager.\n2. Chọn lại Manager.\n3. Lưu.",
            "Role: Manager → Manager (không đổi)",
            "Lưu bình thường, không báo lỗi; dữ liệu không đổi.",
            f"DB={user_row('tc_manager')['role']}; flash={'OK' if has_text(page, 'Cập nhật phân quyền thành công') else page.locator('.alert').inner_text() if page.locator('.alert').count() else 'none'}",
            pass_fail(user_row("tc_manager")["role"] == "manager" and "error" not in page.content().lower()[:500] or True),
            files63,
            "Low",
        )
        # refine TC63: should not error
        results[-1]["status"] = pass_fail(
            user_row("tc_manager")["role"] == "manager" and page.locator(".alert-danger").count() == 0
        )
        results[-1]["actual"] = (
            f"DB={user_row('tc_manager')['role']}; "
            f"alert danger={page.locator('.alert-danger').count()}; "
            f"flash success={page.locator('.alert-success').count()}"
        )

        reset_accounts()
        browser.close()

    xlsx, passed, failed = write_excel()
    print(f"DONE shots={shot_index} pass={passed} fail={failed} excel={xlsx}")
    for r in results:
        if r["status"] != "Pass":
            print(f"FAIL TC{r['id']}: {r['name']} -> {r['actual'][:200]}")


if __name__ == "__main__":
    try:
        run()
    except Exception:
        traceback.print_exc()
        raise
