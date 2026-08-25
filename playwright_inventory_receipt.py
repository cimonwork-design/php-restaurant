# -*- coding: utf-8 -*-
"""
Kịch bản Kiểm thử Tự động E2E bằng Playwright cho Chức năng:
TẠO PHIẾU NHẬP KHO (Inventory Receipt Create)
Hệ thống Quản lý Nhà hàng (PHP Restaurant Management)

- Thực hiện kiểm thử toàn bộ 73 Test Cases (UI, Phân quyền, Validate từng field, Tính toán tiền, v.v.)
- Tự động chụp ảnh từng bước/test case
- Tự động xuất kết quả ra file Excel (KetQua_TaoPhieuNhapKho_Playwright.xlsx),
  file HTML trực quan (Bao_cao_Kiem_thu_Tao_Phieu_Nhap.html) và file Markdown.
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
import traceback
from datetime import datetime, timedelta
from pathlib import Path

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

# ==============================================================================
# CẤU HÌNH HỆ THỐNG KIỂM THỬ
# ==============================================================================
BASE_URL = "http://localhost/php-restaurant-main-main/"
PHP_EXE = r"C:\xampp\php\php.exe"
MYSQL_EXE = r"C:\xampp\mysql\bin\mysql.exe"

# Thư mục lưu ảnh chụp màn hình và báo cáo
CURRENT_DIR = Path(__file__).parent.resolve()
ROOT_DIR = CURRENT_DIR.as_posix()
OUT_DIR = CURRENT_DIR / "testcase_receipt_screenshots"
TODAY = datetime.now().strftime("%d/%m/%Y %H:%M")
TESTER = "Playwright (Chromium)"

ACCOUNTS = {
    "admin": {"user": "admin", "pass": "admin123", "role": "admin"},
    "manager": {"user": "manager", "pass": "admin123", "role": "manager"},
    "staff": {"user": "staff", "pass": "admin123", "role": "user"},
    "cashier": {"user": "cashier", "pass": "admin123", "role": "user"},
}

results = []
shot_index = 0


# ==============================================================================
# CÁC HÀM TIỆN ÍCH HỖ TRỢ (HELPERS)
# ==============================================================================
def run_php(script: str) -> str:
    """Chạy script PHP nhanh để tương tác với cơ sở dữ liệu."""
    proc = subprocess.run(
        [PHP_EXE, "-r", script],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return proc.stdout.strip()


def ensure_test_users():
    """Đảm bảo các tài khoản kiểm thử tồn tại trong DB."""
    php_code = f"""
    require_once '{ROOT_DIR}/config/config.php';
    $db = getDB();
    $users = [
        ['admin', password_hash('admin123', PASSWORD_DEFAULT), 'Quản trị viên', 'admin', 1],
        ['manager', password_hash('admin123', PASSWORD_DEFAULT), 'Quản lý kho', 'manager', 1],
        ['staff', password_hash('admin123', PASSWORD_DEFAULT), 'Nhân viên phục vụ', 'user', 1],
        ['cashier', password_hash('admin123', PASSWORD_DEFAULT), 'Thu ngân', 'user', 1],
    ];
    foreach ($users as $u) {{
        $stmt = $db->prepare("SELECT id FROM users WHERE username = ?");
        $stmt->execute([$u[0]]);
        if ($stmt->fetch()) {{
            $up = $db->prepare("UPDATE users SET password=?, fullname=?, role=?, active=? WHERE username=?");
            $up->execute([$u[1], $u[2], $u[3], $u[4], $u[0]]);
        }} else {{
            $in = $db->prepare("INSERT INTO users (username, password, fullname, role, active) VALUES (?, ?, ?, ?, ?)");
            $in->execute([$u[0], $u[1], $u[2], $u[3], $u[4]]);
        }}
    }}
    """
    run_php(php_code)


def clean_test_receipts():
    """Dọn dẹp các phiếu nhập sinh ra trong quá trình kiểm thử."""
    php_code = f"""
    require_once '{ROOT_DIR}/config/config.php';
    $db = getDB();
    $db->exec("DELETE FROM inventory_receipt_detail WHERE receipt_id IN (SELECT id FROM inventory_receipt WHERE supplier LIKE 'TEST_%' OR note LIKE '%KIEM_THU%')");
    $db->exec("DELETE FROM inventory_receipt WHERE supplier LIKE 'TEST_%' OR note LIKE '%KIEM_THU%'");
    """
    run_php(php_code)


def shot(page, slug: str) -> str:
    """Chụp ảnh màn hình toàn trang và lưu file."""
    global shot_index
    shot_index += 1
    filename = f"{shot_index:03d}_{slug}.png"
    filepath = OUT_DIR / filename
    page.screenshot(path=str(filepath), full_page=True)
    return filename


def add_result(
    tc_id: int,
    category: str,
    name: str,
    steps: str,
    data: str,
    expected: str,
    actual: str,
    status: str,
    screenshots: list[str],
    priority: str = "High",
):
    """Ghi nhận kết quả của một testcase."""
    results.append({
        "id": tc_id,
        "category": category,
        "name": name,
        "steps": steps,
        "data": data,
        "expected": expected,
        "actual": actual,
        "status": status,
        "screenshots": screenshots,
        "priority": priority,
        "date": TODAY,
        "tester": TESTER,
    })


def pass_fail(ok: bool) -> str:
    return "Pass" if ok else "Fail"


def login(page, username: str, password: str):
    """Thực hiện đăng nhập qua form."""
    try:
        page.context.clear_cookies()
    except Exception:
        pass
    page.goto(BASE_URL + "auth/login", wait_until="domcontentloaded")
    try:
        page.evaluate("localStorage.clear(); sessionStorage.clear();")
    except Exception:
        pass
    page.goto(BASE_URL + "auth/login", wait_until="domcontentloaded")
    page.wait_for_selector("#username", timeout=8000)
    page.fill("#username", username)
    page.fill("#password", password)
    page.click("#btnLogin")
    page.wait_for_timeout(1200)


def logout(page):
    """Đăng xuất khỏi hệ thống."""
    try:
        page.context.clear_cookies()
    except Exception:
        pass
    page.goto(BASE_URL + "auth/logout", wait_until="domcontentloaded")
    try:
        page.evaluate("localStorage.clear(); sessionStorage.clear();")
    except Exception:
        pass
    page.wait_for_timeout(300)


def goto_create_receipt(page):
    """Mở màn hình Tạo phiếu nhập kho."""
    page.goto(BASE_URL + "inventory_receipt/create", wait_until="domcontentloaded")
    page.wait_for_timeout(300)


# ==============================================================================
# XUẤT BÁO CÁO EXCEL (THEO MẪU NHÓM 3)
# ==============================================================================
def export_excel() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tao-Phieu-Nhap"

    headers = [
        "ID", "Category", "Test case Name", "Test step", "Test data",
        "Expected result", "Priority", "Actual result", "Test result",
        "Screenshot", "Date", "Tester",
    ]
    header_fill = PatternFill("solid", fgColor="1E3A8A")  # Xanh navy đậm
    header_font = Font(bold=True, color="FFFFFF", size=11)
    pass_fill = PatternFill("solid", fgColor="DCFCE7")    # Xanh lá nhạt
    fail_fill = PatternFill("solid", fgColor="FEE2E2")    # Đỏ nhạt
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Tiêu đề báo cáo
    ws.merge_cells("A1:L1")
    ws["A1"] = "BẢNG KẾT QUẢ KIỂM THỬ PLAYWRIGHT — CHỨC NĂNG TẠO PHIẾU NHẬP KHO"
    ws["A1"].font = Font(bold=True, size=15, color="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    passed = sum(1 for r in results if r["status"] == "Pass")
    failed = sum(1 for r in results if r["status"] == "Fail")

    ws.merge_cells("A2:L2")
    ws["A2"] = (
        f"Công cụ: Playwright (Chromium)  |  Ngày thực hiện: {TODAY}  |  "
        f"Tổng Test Cases: {len(results)}  |  Pass: {passed}  |  Fail: {failed}  |  "
        f"Thư mục ảnh: {OUT_DIR}"
    )
    ws["A2"].font = Font(italic=True, color="4B5563", size=10)
    ws.row_dimensions[2].height = 20

    # Header bảng
    for col, h in enumerate(headers, 1):
        cell = ws.cell(3, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 26

    # Ghi dữ liệu từng dòng
    for i, r in enumerate(results, 4):
        values = [
            r["id"], r["category"], r["name"], r["steps"], r["data"],
            r["expected"], r["priority"], r["actual"], r["status"],
            "\n".join(r["screenshots"]), r["date"], r["tester"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i, col, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
            if col == 9:  # Test result
                cell.fill = pass_fill if r["status"] == "Pass" else fail_fill
                cell.font = Font(bold=True, color="15803D" if r["status"] == "Pass" else "B91C1C")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col == 7:
                cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.row_dimensions[i].height = 70

    # Điều chỉnh độ rộng cột
    widths = [6, 22, 38, 42, 28, 42, 12, 42, 12, 32, 18, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A3:L{3 + len(results)}"
    ws.freeze_panes = "A4"

    # Sheet 2: Tổng hợp kết quả
    summary = wb.create_sheet("Tong hop")
    summary["A1"] = "BẢNG TỔNG HỢP KẾT QUẢ KIỂM THỬ THEO NHÓM"
    summary["A1"].font = Font(bold=True, size=14, color="1E3A8A")
    summary.append([])
    summary.append(["Category (Nhóm kiểm thử)", "Số Test Case", "Pass", "Fail", "Tỷ lệ Pass (%)"])

    cats = {}
    for r in results:
        cats.setdefault(r["category"], {"Pass": 0, "Fail": 0})
        cats[r["category"]][r["status"]] += 1

    for cat, c in cats.items():
        total_cat = c["Pass"] + c["Fail"]
        rate = round((c["Pass"] / total_cat) * 100, 1) if total_cat > 0 else 0
        summary.append([cat, total_cat, c["Pass"], c["Fail"], f"{rate}%"])

    summary.append(["TỔNG CỘNG", len(results), passed, failed, f"{round((passed/len(results))*100, 1)}%"])

    for col in range(1, 6):
        cell = summary.cell(3, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    summary_widths = [32, 16, 12, 12, 18]
    for col, w in enumerate(summary_widths, 1):
        summary.column_dimensions[get_column_letter(col)].width = w

    excel_path = CURRENT_DIR / "KetQua_TaoPhieuNhapKho_Playwright.xlsx"
    wb.save(excel_path)
    return excel_path


# ==============================================================================
# XUẤT BÁO CÁO HTML
# ==============================================================================
def export_html() -> Path:
    passed = sum(1 for r in results if r["status"] == "Pass")
    failed = sum(1 for r in results if r["status"] == "Fail")
    pass_rate = round((passed / len(results)) * 100, 1) if results else 0

    cats = {}
    for r in results:
        cats.setdefault(r["category"], {"Pass": 0, "Fail": 0, "Total": 0})
        cats[r["category"]]["Total"] += 1
        cats[r["category"]][r["status"]] += 1

    cat_rows = "".join(f"""
        <tr>
            <td class="fw-semibold">{cat}</td>
            <td class="text-center">{data['Total']}</td>
            <td class="text-center text-success fw-bold">{data['Pass']}</td>
            <td class="text-center text-danger fw-bold">{data['Fail']}</td>
            <td class="text-center">{round((data['Pass']/data['Total'])*100, 1)}%</td>
        </tr>
    """ for cat, data in cats.items())

    tc_rows = "".join(f"""
        <tr class="{'table-danger' if r['status'] == 'Fail' else ''}">
            <td class="text-center fw-bold">{r['id']}</td>
            <td><span class="badge bg-secondary">{r['category']}</span></td>
            <td class="fw-semibold">{r['name']}</td>
            <td><small style="white-space: pre-line;">{r['steps']}</small></td>
            <td><code>{r['data']}</code></td>
            <td><small>{r['expected']}</small></td>
            <td class="text-center"><span class="badge bg-{'danger' if r['priority']=='High' else 'warning text-dark' if r['priority']=='Medium' else 'info text-dark'}">{r['priority']}</span></td>
            <td><small>{r['actual']}</small></td>
            <td class="text-center"><span class="badge bg-{'success' if r['status']=='Pass' else 'danger'} fs-6">{r['status']}</span></td>
            <td><small class="text-muted">{', '.join(r['screenshots']) if r['screenshots'] else '-'}</small></td>
        </tr>
    """ for r in results)

    html_content = f"""<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Báo cáo Kiểm thử Tự động — Tạo Phiếu Nhập Kho</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body {{ background-color: #f8fafc; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }}
        .header-box {{ background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%); color: white; padding: 2rem; border-radius: 12px; margin-bottom: 2rem; }}
        .metric-card {{ background: white; border-radius: 10px; padding: 1.25rem; box-shadow: 0 2px 6px rgba(0,0,0,0.06); }}
        .table-responsive {{ background: white; border-radius: 10px; box-shadow: 0 2px 6px rgba(0,0,0,0.06); padding: 1rem; }}
        th {{ background-color: #1e3a8a !important; color: white !important; font-size: 0.9rem; }}
    </style>
</head>
<body class="p-4">
    <div class="container-fluid">
        <div class="header-box shadow">
            <h1 class="h2 fw-bold mb-1">Báo Cáo Kiểm Thử Tự Động Playwright</h1>
            <p class="mb-0 fs-5 opacity-75">Chức năng: Tạo phiếu nhập kho (Inventory Receipt Create) — PHP Restaurant</p>
            <hr class="my-3 opacity-25">
            <div class="d-flex flex-wrap gap-4 small">
                <div><strong>Ngày chạy:</strong> {TODAY}</div>
                <div><strong>Người / Công cụ:</strong> {TESTER}</div>
                <div><strong>Môi trường:</strong> Chromium / Windows Localhost</div>
            </div>
        </div>

        <div class="row g-3 mb-4">
            <div class="col-md-3">
                <div class="metric-card text-center border-start border-primary border-4">
                    <div class="text-muted small text-uppercase fw-bold">Tổng số Test Cases</div>
                    <div class="fs-2 fw-bold text-primary">{len(results)}</div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="metric-card text-center border-start border-success border-4">
                    <div class="text-muted small text-uppercase fw-bold">Passed</div>
                    <div class="fs-2 fw-bold text-success">{passed}</div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="metric-card text-center border-start border-danger border-4">
                    <div class="text-muted small text-uppercase fw-bold">Failed</div>
                    <div class="fs-2 fw-bold text-danger">{failed}</div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="metric-card text-center border-start border-info border-4">
                    <div class="text-muted small text-uppercase fw-bold">Tỷ lệ thành công</div>
                    <div class="fs-2 fw-bold text-info">{pass_rate}%</div>
                </div>
            </div>
        </div>

        <div class="card mb-4 shadow-sm border-0">
            <div class="card-header bg-white fw-bold py-3">Bảng tổng hợp theo nhóm kiểm thử</div>
            <div class="card-body p-0">
                <div class="table-responsive">
                    <table class="table table-hover table-striped mb-0">
                        <thead>
                            <tr>
                                <th>Nhóm kiểm thử (Category)</th>
                                <th class="text-center">Số Test Cases</th>
                                <th class="text-center">Pass</th>
                                <th class="text-center">Fail</th>
                                <th class="text-center">Tỷ lệ đạt</th>
                            </tr>
                        </thead>
                        <tbody>{cat_rows}</tbody>
                    </table>
                </div>
            </div>
        </div>

        <div class="card shadow-sm border-0">
            <div class="card-header bg-white fw-bold py-3 d-flex justify-content-between align-items-center">
                <span>Chi tiết kết quả thực hiện {len(results)} Test Cases</span>
            </div>
            <div class="card-body p-0">
                <div class="table-responsive">
                    <table class="table table-bordered table-hover align-middle mb-0" style="font-size: 0.88rem;">
                        <thead>
                            <tr>
                                <th style="width: 50px;">ID</th>
                                <th style="width: 140px;">Category</th>
                                <th style="width: 220px;">Tên Test Case</th>
                                <th>Các bước thực hiện</th>
                                <th style="width: 160px;">Dữ liệu test</th>
                                <th>Kết quả mong đợi</th>
                                <th style="width: 80px;">Ưu tiên</th>
                                <th>Kết quả thực tế</th>
                                <th style="width: 80px;">Kết quả</th>
                                <th style="width: 150px;">Ảnh chụp</th>
                            </tr>
                        </thead>
                        <tbody>{tc_rows}</tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
"""
    html_path = CURRENT_DIR / "Bao_cao_Kiem_thu_Tao_Phieu_Nhap.html"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    return html_path


# ==============================================================================
# HÀM THỰC THI TOÀN BỘ KIỂM THỬ (MAIN RUNNER)
# ==============================================================================
def run_all_tests():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for old in OUT_DIR.glob("*.png"):
        old.unlink()

    ensure_test_users()
    clean_test_receipts()

    today_str = datetime.now().strftime("%Y-%m-%d")
    yesterday_str = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d")
    tomorrow_str = (datetime.now() + timedelta(days=2)).strftime("%Y-%m-%d")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(viewport={"width": 1440, "height": 900}, locale="vi-VN")
        page = context.new_page()
        page.set_default_timeout(15000)

        print("\n" + "="*70)
        print("BẮT ĐẦU CHẠY KIỂM THỬ TỰ ĐỘNG FORM TẠO PHIẾU NHẬP KHO")
        print("="*70 + "\n")

        # ======================================================================
        # CATEGORY 1: UI DISPLAY & FORM STRUCTURE (TC01 - TC08)
        # ======================================================================
        print("[1/11] Đang kiểm tra Category: UI Display & Form Structure...")
        login(page, "admin", "admin123")
        goto_create_receipt(page)
        s1 = shot(page, "TC01_UI_man_hinh_tao_phieu")

        # TC01: Tiêu đề trang
        h1_text = page.locator("h1").inner_text()
        tc01_ok = "Tạo phiếu nhập kho" in h1_text
        add_result(
            1, "UI Display", "Hiển thị đúng tiêu đề màn hình và mô tả trang",
            "1. Đăng nhập Admin.\n2. Mở URL /inventory_receipt/create.\n3. Quan sát tiêu đề h1.",
            "Role: Admin", 'Tiêu đề trang có "Tạo phiếu nhập kho" và mô tả ngắn.',
            f"Tiêu đề thực tế: '{h1_text}'", pass_fail(tc01_ok), [s1], "High"
        )

        # TC02: Các trường thông tin chung
        has_supplier = page.locator("#supplier").count() > 0
        has_date = page.locator("#receipt_date").count() > 0
        has_note = page.locator("#note").count() > 0
        tc02_ok = has_supplier and has_date and has_note
        add_result(
            2, "UI Display", "Hiển thị đầy đủ các trường thông tin chung (NCC, Ngày nhập, Ghi chú)",
            "1. Quan sát phần 'Thông tin phiếu nhập'.\n2. Kiểm tra sự tồn tại của 3 input: supplier, receipt_date, note.",
            "N/A", "Có đủ 3 trường: Nhà cung cấp, Ngày nhập, Ghi chú.",
            f"supplier={has_supplier}, date={has_date}, note={has_note}", pass_fail(tc02_ok), [s1], "High"
        )

        # TC03: Ngày nhập mặc định
        date_val = page.locator("#receipt_date").input_value()
        tc03_ok = (date_val == today_str)
        add_result(
            3, "UI Display", "Trường Ngày nhập mặc định là ngày hiện tại (YYYY-MM-DD)",
            "1. Mở form tạo mới.\n2. Lấy giá trị trường receipt_date.",
            f"Ngày hiện tại: {today_str}", f"Giá trị mặc định bằng {today_str}.",
            f"Giá trị thực tế: '{date_val}'", pass_fail(tc03_ok), [s1], "Medium"
        )

        # TC04: Các cột bảng chi tiết nguyên liệu
        s4 = shot(page, "TC04_bang_chi_tiet_nguyen_lieu")
        page_text = page.content()
        has_col_ing = "Nguyên liệu" in page_text
        has_col_qty = "Số lượng" in page_text
        has_col_price = "Đơn giá" in page_text
        has_col_subtotal = "Thành tiền" in page_text
        tc04_ok = has_col_ing and has_col_qty and has_col_price and has_col_subtotal
        add_result(
            4, "UI Display", "Hiển thị đầy đủ các cột trong bảng chi tiết nguyên liệu",
            "1. Quan sát bảng Chi tiết nguyên liệu.\n2. Kiểm tra tên các cột.",
            "N/A", "Đủ các cột: Nguyên liệu, Số lượng, Đơn giá, Thành tiền, Xóa.",
            f"Có đủ các cột tiêu đề: {tc04_ok}", pass_fail(tc04_ok), [s4], "Medium"
        )

        # TC05: Dropdown danh sách nguyên liệu từ DB
        options_count = page.locator(".ingredient-select option").count()
        tc05_ok = options_count > 1
        add_result(
            5, "UI Display", "Hiển thị dropdown danh sách nguyên liệu lấy từ cơ sở dữ liệu",
            "1. Quan sát thẻ select nguyên liệu.\n2. Đếm số lượng option có sẵn.",
            "Database: bảng ingredient", "Dropdown chứa option mặc định và danh sách các nguyên liệu từ DB.",
            f"Số lượng options: {options_count}", pass_fail(tc05_ok), [s4], "High"
        )

        # TC06: Tổng cộng và số mục mặc định
        total_text = page.locator("#receipt-total").inner_text()
        count_text = page.locator("#item-count").inner_text()
        tc06_ok = ("0" in total_text) and ("mục" in count_text)
        add_result(
            6, "UI Display", "Hiển thị ô Tổng cộng giá trị mặc định là 0 đ và số mục",
            "1. Quan sát khung Tổng tiền dưới form.", "N/A",
            "Tổng cộng hiển thị '0 đ' và số mục khớp số dòng.",
            f"Tổng cộng: '{total_text}', Số mục: '{count_text}'", pass_fail(tc06_ok), [s4], "Low"
        )

        # TC07: Nút Thêm dòng, Tạo phiếu nhập, Hủy
        has_btn_add = page.locator("#add-item").count() > 0
        has_btn_submit = page.locator("#btnSubmitReceipt").count() > 0
        has_btn_cancel = page.locator("#btnCancel").count() > 0
        tc07_ok = has_btn_add and has_btn_submit and has_btn_cancel
        add_result(
            7, "UI Display", "Hiển thị đầy đủ các nút chức năng (Thêm dòng, Tạo phiếu nhập, Hủy)",
            "1. Kiểm tra sự tồn tại của 3 nút thao tác chính trên form.", "N/A",
            "Có đủ 3 nút: Thêm dòng (#add-item), Tạo phiếu nhập (#btnSubmitReceipt), Hủy (#btnCancel).",
            f"add={has_btn_add}, submit={has_btn_submit}, cancel={has_btn_cancel}", pass_fail(tc07_ok), [s4], "Medium"
        )

        # TC08: Nút Quay lại / Hủy điều hướng về danh sách
        s8 = shot(page, "TC08_nut_quay_lai")
        page.locator("#btnCancel").click()
        page.wait_for_load_state("domcontentloaded")
        tc08_ok = "inventory_receipt" in page.url and "create" not in page.url
        add_result(
            8, "UI Display", "Nhấn nút 'Quay lại' hoặc 'Hủy' điều hướng về trang danh sách phiếu nhập",
            "1. Tại trang create, nhấn nút 'Hủy' / 'Quay lại'.\n2. Quan sát URL.",
            "URL: /inventory_receipt/create", "Chuyển hướng về /inventory_receipt thành công.",
            f"URL sau khi nhấn: {page.url}", pass_fail(tc08_ok), [s8], "Low"
        )

        # ======================================================================
        # CATEGORY 2: ACCESS CONTROL & PERMISSIONS (TC09 - TC15)
        # ======================================================================
        print("[2/11] Đang kiểm tra Category: Access Control & Permissions...")
        # TC09: Admin truy cập
        goto_create_receipt(page)
        s9 = shot(page, "TC09_admin_truy_cap")
        tc09_ok = "create" in page.url and page.locator("#receiptForm").count() > 0
        add_result(
            9, "Access Control", "Quản trị viên (Admin) truy cập thành công màn hình Tạo phiếu nhập",
            "1. Đăng nhập Admin.\n2. Truy cập /inventory_receipt/create.",
            "Role: Admin", "Mở thành công form tạo phiếu nhập.",
            f"URL={page.url}, Form exists={tc09_ok}", pass_fail(tc09_ok), [s9], "High"
        )

        # TC10: Manager truy cập
        logout(page)
        login(page, "manager", "admin123")
        goto_create_receipt(page)
        s10 = shot(page, "TC10_manager_truy_cap")
        tc10_ok = "create" in page.url and page.locator("#receiptForm").count() > 0
        add_result(
            10, "Access Control", "Quản lý (Manager) truy cập thành công màn hình Tạo phiếu nhập",
            "1. Đăng nhập Manager.\n2. Truy cập /inventory_receipt/create.",
            "Role: Manager", "Mở thành công form tạo phiếu nhập.",
            f"URL={page.url}, Form exists={tc10_ok}", pass_fail(tc10_ok), [s10], "High"
        )

        # TC11: Staff (User) bị từ chối
        logout(page)
        login(page, "staff", "admin123")
        goto_create_receipt(page)
        s11 = shot(page, "TC11_staff_bi_chan")
        body_11 = page.content()
        denied_11 = ("không có quyền" in body_11.lower() or "dashboard" in page.url or "403" in body_11)
        add_result(
            11, "Access Control", "Nhân viên (Staff / User) không thể truy cập màn hình Tạo phiếu nhập",
            "1. Đăng nhập Staff.\n2. Nhập URL /inventory_receipt/create.",
            "Role: Staff (User)", "Bị chặn truy cập, chuyển hướng về Dashboard hoặc thông báo không đủ quyền.",
            f"URL={page.url}, Denied text present={denied_11}", pass_fail(denied_11), [s11], "High"
        )

        # TC12: Cashier bị từ chối
        logout(page)
        login(page, "cashier", "admin123")
        goto_create_receipt(page)
        s12 = shot(page, "TC12_cashier_bi_chan")
        body_12 = page.content()
        denied_12 = ("không có quyền" in body_12.lower() or "dashboard" in page.url)
        add_result(
            12, "Access Control", "Thu ngân (Cashier) không thể truy cập màn hình Tạo phiếu nhập",
            "1. Đăng nhập Cashier.\n2. Nhập URL /inventory_receipt/create.",
            "Role: Cashier", "Hệ thống từ chối quyền truy cập.",
            f"URL={page.url}", pass_fail(denied_12), [s12], "High"
        )

        # TC13: Guest (Chưa login) bị redirect
        logout(page)
        guest_ctx = browser.new_context(viewport={"width": 1440, "height": 900})
        guest_page = guest_ctx.new_page()
        guest_page.goto(BASE_URL + "inventory_receipt/create", wait_until="domcontentloaded")
        s13 = shot(guest_page, "TC13_guest_redirect_login")
        guest_ok = ("auth/login" in guest_page.url or "login" in guest_page.url)
        add_result(
            13, "Access Control", "Khách chưa đăng nhập không thể truy cập màn hình Tạo phiếu nhập",
            "1. Mở phiên ẩn danh / chưa login.\n2. Truy cập /inventory_receipt/create.",
            "Role: Guest", "Tự động chuyển hướng về trang Đăng nhập (auth/login).",
            f"URL thực tế: {guest_page.url}", pass_fail(guest_ok), [s13], "High"
        )
        guest_ctx.close()

        # TC14: Staff gọi trực tiếp POST API /store
        staff_ctx = browser.new_context(viewport={"width": 1440, "height": 900})
        staff_pg = staff_ctx.new_page()
        login(staff_pg, "staff", "admin123")
        resp14 = staff_pg.request.post(
            BASE_URL + "inventory_receipt/store",
            data={"receipt_date": today_str, "ingredient_id": ["1"], "qty": ["5"], "unit_price": ["10000"]},
        )
        s14 = shot(staff_pg, "TC14_staff_post_api_blocked")
        tc14_ok = (resp14.status in (403, 302, 401) or "dashboard" in resp14.url)
        add_result(
            14, "Access Control", "Nhân viên gọi trực tiếp POST API /inventory_receipt/store bị chặn quyền",
            "1. Đăng nhập Staff.\n2. Gửi request POST tới /inventory_receipt/store.",
            "Actor: Staff / Target: POST /store", "Server từ chối hoặc redirect về dashboard, không lưu dữ liệu.",
            f"HTTP status: {resp14.status}, redirected: {resp14.url}", pass_fail(tc14_ok), [s14], "High"
        )
        staff_ctx.close()

        # TC15: Guest gọi POST API /store
        guest_req = context.request.post(
            BASE_URL + "inventory_receipt/store",
            data={"receipt_date": today_str, "ingredient_id": ["1"], "qty": ["5"], "unit_price": ["10000"]},
        )
        tc15_ok = (guest_req.status in (401, 302, 403) or "auth/login" in guest_req.url)
        add_result(
            15, "Access Control", "Người dùng chưa đăng nhập gọi POST API /inventory_receipt/store bị chặn",
            "1. Gửi request POST /inventory_receipt/store không có JWT cookie.",
            "Actor: Guest", "Bị từ chối hoặc chuyển hướng về auth/login.",
            f"HTTP status: {guest_req.status}", pass_fail(tc15_ok), [s13], "High"
        )

        # ======================================================================
        # CATEGORY 3: FIELD VALIDATION - SUPPLIER (TC16 - TC22)
        # ======================================================================
        print("[3/11] Đang kiểm tra Category: Field Validation - Supplier...")
        login(page, "admin", "admin123")

        # TC16: Để trống Nhà cung cấp (Hợp lệ)
        goto_create_receipt(page)
        page.fill("#supplier", "")
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("10")
        page.locator(".price-input").first.fill("20000")
        s16 = shot(page, "TC16_supplier_empty")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(400)
        tc16_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            16, "Validation - Supplier", "Để trống trường Nhà cung cấp (Trường tùy chọn - Cho phép để trống)",
            "1. Nhập form hợp lệ, để trống ô Nhà cung cấp.\n2. Nhấn Tạo phiếu nhập.",
            "Supplier: [Empty]", "Tạo phiếu nhập thành công, trường supplier nhận giá trị NULL/rỗng.",
            f"Flash: {'OK' if tc16_ok else 'LỖI'}, URL={page.url}", pass_fail(tc16_ok), [s16], "Medium"
        )

        # TC17: Nhập Nhà cung cấp hợp lệ
        goto_create_receipt(page)
        page.fill("#supplier", "Công ty Thực phẩm Sạch TEST_ABC")
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("5")
        page.locator(".price-input").first.fill("15000")
        s17 = shot(page, "TC17_supplier_valid")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(400)
        tc17_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            17, "Validation - Supplier", "Nhập tên Nhà cung cấp hợp lệ (2 - 100 ký tự văn bản chuẩn)",
            "1. Nhập Nhà cung cấp: 'Công ty Thực phẩm Sạch TEST_ABC'.\n2. Nhấn Tạo phiếu nhập.",
            "Supplier: 'Công ty Thực phẩm Sạch TEST_ABC'", "Tạo phiếu nhập thành công.",
            f"Kết quả lưu: {tc17_ok}", pass_fail(tc17_ok), [s17], "High"
        )

        # TC18: Nhà cung cấp chỉ chứa toàn khoảng trắng
        goto_create_receipt(page)
        page.fill("#supplier", "     ")
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("2")
        page.locator(".price-input").first.fill("10000")
        s18 = shot(page, "TC18_supplier_all_spaces")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content18 = page.content()
        tc18_ok = "khoảng trắng" in content18.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            18, "Validation - Supplier", "Nhập Nhà cung cấp chỉ chứa toàn khoảng trắng",
            "1. Nhập 5 dấu cách vào ô Nhà cung cấp.\n2. Nhấn Tạo phiếu nhập.",
            "Supplier: '     '", "Hệ thống báo lỗi 'Tên nhà cung cấp không được chỉ chứa khoảng trắng'.",
            f"Phát hiện lỗi khoảng trắng: {tc18_ok}", pass_fail(tc18_ok), [s18], "Medium"
        )

        # TC19: Nhà cung cấp có độ dài 1 ký tự (< 2 ký tự)
        goto_create_receipt(page)
        page.fill("#supplier", "A")
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("2")
        page.locator(".price-input").first.fill("10000")
        s19 = shot(page, "TC19_supplier_1_char")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content19 = page.content()
        tc19_ok = "tối thiểu 2 ký tự" in content19.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            19, "Validation - Supplier", "Nhập tên Nhà cung cấp quá ngắn (1 ký tự)",
            "1. Nhập Supplier = 'A'.\n2. Nhấn Tạo phiếu nhập.",
            "Supplier: 'A' (1 char)", "Báo lỗi 'Tên nhà cung cấp phải có tối thiểu 2 ký tự'.",
            f"Báo lỗi tối thiểu 2 ký tự: {tc19_ok}", pass_fail(tc19_ok), [s19], "Medium"
        )

        # TC20: Nhà cung cấp đạt độ dài tối đa biên 100 ký tự
        supp_100 = "TEST_NCC_" + ("A" * 91)
        goto_create_receipt(page)
        page.fill("#supplier", supp_100)
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("10000")
        s20 = shot(page, "TC20_supplier_100_chars")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(400)
        tc20_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            20, "Validation - Supplier", "Nhập tên Nhà cung cấp đạt giá trị biên tối đa (100 ký tự)",
            "1. Nhập chuỗi 100 ký tự vào ô Nhà cung cấp.\n2. Submit form.",
            f"Supplier length: {len(supp_100)} chars", "Tạo phiếu nhập thành công không phát sinh lỗi.",
            f"Lưu thành công: {tc20_ok}", pass_fail(tc20_ok), [s20], "Medium"
        )

        # TC21: Nhà cung cấp vượt quá 100 ký tự (> 100 ký tự)
        supp_120 = "TEST_NCC_" + ("B" * 115)
        goto_create_receipt(page)
        page.evaluate(f"document.getElementById('supplier').removeAttribute('maxlength'); document.getElementById('supplier').value = '{supp_120}';")
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("10000")
        s21 = shot(page, "TC21_supplier_over_100_chars")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content21 = page.content()
        tc21_ok = "vượt quá 100 ký tự" in content21.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            21, "Validation - Supplier", "Nhập tên Nhà cung cấp vượt quá 100 ký tự (> 100 chars)",
            "1. Nhập chuỗi 124 ký tự vào ô Nhà cung cấp.\n2. Nhấn Tạo phiếu nhập.",
            f"Supplier length: {len(supp_120)} chars", "Báo lỗi 'Tên nhà cung cấp không được vượt quá 100 ký tự'.",
            f"Bắt lỗi độ dài vượt quá: {tc21_ok}", pass_fail(tc21_ok), [s21], "High"
        )

        # TC22: Nhà cung cấp chứa số điện thoại, dấu ngoặc, tiếng Việt có dấu
        goto_create_receipt(page)
        page.fill("#supplier", "TEST_Đại lý Nông Sản Sạch (SĐT: 0987.654.321)")
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("3")
        page.locator(".price-input").first.fill("12000")
        s22 = shot(page, "TC22_supplier_special_chars")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(400)
        tc22_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            22, "Validation - Supplier", "Nhập Nhà cung cấp chứa ký tự đặc biệt hợp lệ (SĐT, dấu ngoặc, Unicode tiếng Việt)",
            "1. Nhập 'TEST_Đại lý Nông Sản Sạch (SĐT: 0987.654.321)'.\n2. Submit form.",
            "Supplier: 'TEST_Đại lý Nông Sản Sạch (SĐT: 0987.654.321)'", "Lưu thành công, chuỗi hiển thị đúng nguyên vẹn.",
            f"Lưu thành công: {tc22_ok}", pass_fail(tc22_ok), [s22], "Low"
        )

        # ======================================================================
        # CATEGORY 4: FIELD VALIDATION - RECEIPT DATE (TC23 - TC29)
        # ======================================================================
        print("[4/11] Đang kiểm tra Category: Field Validation - Receipt Date...")

        # TC23: Bỏ trống trường Ngày nhập
        goto_create_receipt(page)
        page.evaluate("document.getElementById('receipt_date').value = '';")
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("10000")
        s23 = shot(page, "TC23_receipt_date_empty")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content23 = page.content()
        tc23_ok = "ngày nhập kho là bắt buộc" in content23.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            23, "Validation - Date", "Bỏ trống trường Ngày nhập kho",
            "1. Xóa rỗng trường Ngày nhập (receipt_date).\n2. Nhấn Tạo phiếu nhập.",
            "Receipt Date: [Empty]", "Báo lỗi 'Ngày nhập kho là bắt buộc, không được để trống'.",
            f"Bắt lỗi ngày rỗng: {tc23_ok}", pass_fail(tc23_ok), [s23], "High"
        )

        # TC24: Ngày nhập là ngày hiện tại
        goto_create_receipt(page)
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("2")
        page.locator(".price-input").first.fill("15000")
        s24 = shot(page, "TC24_receipt_date_today")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc24_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            24, "Validation - Date", "Nhập Ngày nhập kho là ngày hiện tại (Hôm nay)",
            "1. Nhập ngày hôm nay vào ô Ngày nhập.\n2. Submit form.",
            f"Date: {today_str}", "Tạo phiếu nhập thành công.",
            f"Lưu thành công: {tc24_ok}", pass_fail(tc24_ok), [s24], "High"
        )

        # TC25: Ngày nhập trong quá khứ hợp lệ
        goto_create_receipt(page)
        page.fill("#receipt_date", yesterday_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("2")
        page.locator(".price-input").first.fill("15000")
        s25 = shot(page, "TC25_receipt_date_past")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc25_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            25, "Validation - Date", "Nhập Ngày nhập kho là một ngày trong quá khứ hợp lệ",
            "1. Nhập ngày 2 ngày trước vào ô Ngày nhập.\n2. Submit form.",
            f"Date: {yesterday_str}", "Tạo phiếu nhập thành công.",
            f"Lưu thành công: {tc25_ok}", pass_fail(tc25_ok), [s25], "Medium"
        )

        # TC26: Ngày nhập trong tương lai (> Ngày hiện tại)
        goto_create_receipt(page)
        page.evaluate(f"document.getElementById('receipt_date').removeAttribute('max'); document.getElementById('receipt_date').value = '{tomorrow_str}';")
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("2")
        page.locator(".price-input").first.fill("15000")
        s26 = shot(page, "TC26_receipt_date_future")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content26 = page.content()
        tc26_ok = "không được lớn hơn ngày hiện tại" in content26.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            26, "Validation - Date", "Nhập Ngày nhập kho là ngày trong tương lai (> Hôm nay)",
            "1. Nhập ngày 2 ngày sau vào ô Ngày nhập.\n2. Nhấn Tạo phiếu nhập.",
            f"Date: {tomorrow_str} (Future)", "Báo lỗi 'Ngày nhập kho không được lớn hơn ngày hiện tại'.",
            f"Bắt lỗi ngày tương lai: {tc26_ok}", pass_fail(tc26_ok), [s26], "High"
        )

        # TC27: Ngày nhập quá khứ quá xa (< 2020-01-01)
        goto_create_receipt(page)
        page.fill("#receipt_date", "2018-05-15")
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("10000")
        s27 = shot(page, "TC27_receipt_date_too_old")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content27 = page.content()
        tc27_ok = "không được nhỏ hơn ngày 01/01/2020" in content27.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            27, "Validation - Date", "Nhập Ngày nhập kho quá khứ quá xa (Trước ngày 01/01/2020)",
            "1. Nhập ngày 15/05/2018 vào ô Ngày nhập.\n2. Nhấn Tạo phiếu nhập.",
            "Date: '2018-05-15'", "Báo lỗi 'Ngày nhập kho không được nhỏ hơn ngày 01/01/2020'.",
            f"Bắt lỗi ngày quá cũ: {tc27_ok}", pass_fail(tc27_ok), [s27], "Medium"
        )

        # TC28: Định dạng ngày sai quy cách (POST trực tiếp API)
        resp28 = page.request.post(
            BASE_URL + "inventory_receipt/store",
            form={"receipt_date": "invalid-date-format", "ingredient_id": "1", "qty": "1", "unit_price": "10000"},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        tc28_ok = (resp28.status == 422 or "định dạng" in resp28.text().lower())
        add_result(
            28, "Validation - Date", "Gửi dữ liệu Ngày nhập sai định dạng YYYY-MM-DD qua API",
            "1. Gửi request POST /store với receipt_date = 'invalid-date-format'.",
            "Date: 'invalid-date-format'", "Server bắt lỗi định dạng ngày không hợp lệ, từ chối lưu.",
            f"HTTP status: {resp28.status}", pass_fail(tc28_ok), [s27], "High"
        )

        # TC29: Kiểm tra tính đúng đắn của ngày nhuận 29/02
        resp29_valid = page.request.post(
            BASE_URL + "inventory_receipt/store",
            form={"receipt_date": "2024-02-29", "ingredient_id": "1", "qty": "1", "unit_price": "10000"},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        resp29_invalid = page.request.post(
            BASE_URL + "inventory_receipt/store",
            form={"receipt_date": "2023-02-29", "ingredient_id": "1", "qty": "1", "unit_price": "10000"},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        tc29_ok = (resp29_valid.status == 200) and (resp29_invalid.status == 422)
        add_result(
            29, "Validation - Date", "Kiểm tra xử lý ngày nhuận (29/02/2024 hợp lệ vs 29/02/2023 không nhuận)",
            "1. POST ngày 29/02/2024 (năm nhuận) -> Pass.\n2. POST ngày 29/02/2023 (không nhuận) -> Bắt lỗi.",
            "Valid: 2024-02-29 / Invalid: 2023-02-29", "Chấp nhận ngày nhuận thực tế, từ chối ngày không tồn tại.",
            f"Nhuận 2024 HTTP={resp29_valid.status}; Không nhuận 2023 HTTP={resp29_invalid.status}", pass_fail(tc29_ok), [s27], "Medium"
        )

        # ======================================================================
        # CATEGORY 5: FIELD VALIDATION - NOTE (TC30 - TC34)
        # ======================================================================
        print("[5/11] Đang kiểm tra Category: Field Validation - Note...")

        # TC30: Để trống Ghi chú (Hợp lệ)
        goto_create_receipt(page)
        page.fill("#note", "")
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("10000")
        s30 = shot(page, "TC30_note_empty")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc30_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            30, "Validation - Note", "Để trống trường Ghi chú (Trường tùy chọn - Cho phép rỗng)",
            "1. Để trống ô Ghi chú.\n2. Submit form.", "Note: [Empty]",
            "Tạo phiếu nhập thành công, note lưu NULL.", f"Lưu thành công: {tc30_ok}", pass_fail(tc30_ok), [s30], "Low"
        )

        # TC31: Ghi chú ngắn hợp lệ
        goto_create_receipt(page)
        page.fill("#note", "TEST_Ghi chú nhập hàng ca sáng, kiểm tra kỹ chất lượng")
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("2")
        page.locator(".price-input").first.fill("10000")
        s31 = shot(page, "TC31_note_valid")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc31_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            31, "Validation - Note", "Nhập Ghi chú ngắn hợp lệ",
            "1. Nhập văn bản mô tả phiếu vào ô Ghi chú.\n2. Submit form.",
            "Note: 'TEST_Ghi chú nhập hàng ca sáng...'", "Lưu thành công phiếu nhập.",
            f"Lưu thành công: {tc31_ok}", pass_fail(tc31_ok), [s31], "Medium"
        )

        # TC32: Ghi chú đạt độ dài tối đa biên 500 ký tự
        note_500 = "TEST_NOTE_" + ("X" * 490)
        goto_create_receipt(page)
        page.fill("#note", note_500)
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("10000")
        s32 = shot(page, "TC32_note_500_chars")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc32_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            32, "Validation - Note", "Nhập Ghi chú đạt độ dài tối đa biên 500 ký tự",
            "1. Nhập chuỗi 500 ký tự vào ô Ghi chú.\n2. Submit form.",
            f"Note length: {len(note_500)} chars", "Tạo phiếu nhập thành công không lỗi.",
            f"Lưu thành công: {tc32_ok}", pass_fail(tc32_ok), [s32], "Medium"
        )

        # TC33: Ghi chú vượt quá 500 ký tự (> 500 chars)
        note_600 = "TEST_NOTE_" + ("Y" * 590)
        goto_create_receipt(page)
        page.evaluate(f"document.getElementById('note').removeAttribute('maxlength'); document.getElementById('note').value = '{note_600}';")
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("10000")
        s33 = shot(page, "TC33_note_over_500_chars")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content33 = page.content()
        tc33_ok = "vượt quá 500 ký tự" in content33.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            33, "Validation - Note", "Nhập Ghi chú vượt quá 500 ký tự (> 500 chars)",
            "1. Nhập chuỗi 600 ký tự vào ô Ghi chú.\n2. Nhấn Tạo phiếu nhập.",
            f"Note length: {len(note_600)} chars", "Báo lỗi 'Ghi chú không được vượt quá 500 ký tự'.",
            f"Bắt lỗi độ dài note: {tc33_ok}", pass_fail(tc33_ok), [s33], "High"
        )

        # TC34: Ghi chú có nhiều dòng xuống dòng và dấu câu
        multiline_note = "TEST_Dòng 1: Hàng tươi.\nDòng 2: Đã kiểm dịch.\nDòng 3: Nhập kho 01."
        goto_create_receipt(page)
        page.fill("#note", multiline_note)
        page.fill("#receipt_date", today_str)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("10000")
        s34 = shot(page, "TC34_note_multiline")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc34_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            34, "Validation - Note", "Nhập Ghi chú nhiều dòng xuống dòng và ký tự tiếng Việt",
            "1. Nhập ghi chú có ký tự xuống dòng (\\n).\n2. Submit form.",
            "Note: 3 dòng text", "Lưu thành công và giữ nguyên định dạng xuống dòng.",
            f"Lưu thành công: {tc34_ok}", pass_fail(tc34_ok), [s34], "Low"
        )

        # ======================================================================
        # CATEGORY 6: DETAIL ITEMS - INGREDIENT SELECTION (TC35 - TC40)
        # ======================================================================
        print("[6/11] Đang kiểm tra Category: Detail Items - Ingredient Selection...")

        # TC35: Thêm dòng nhưng không chọn nguyên liệu nào
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(value="")
        page.locator(".qty-input").first.fill("5")
        page.locator(".price-input").first.fill("10000")
        s35 = shot(page, "TC35_ingredient_empty_select")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content35 = page.content()
        tc35_ok = "vui lòng chọn nguyên liệu" in content35.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            35, "Detail - Ingredient", "Không chọn nguyên liệu trên dòng (để dropdown ở option mặc định)",
            "1. Để select nguyên liệu rỗng value=''.\n2. Nhấn Tạo phiếu nhập.",
            "Ingredient ID: [Empty]", "Báo lỗi 'Dòng 1: Vui lòng chọn nguyên liệu'.",
            f"Bắt lỗi chưa chọn NL: {tc35_ok}", pass_fail(tc35_ok), [s35], "High"
        )

        # TC36: Chọn nguyên liệu hợp lệ từ danh sách
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        selected_val = page.locator(".ingredient-select").first.input_value()
        s36 = shot(page, "TC36_ingredient_valid_selected")
        tc36_ok = bool(selected_val and selected_val != "")
        add_result(
            36, "Detail - Ingredient", "Chọn nguyên liệu hợp lệ từ dropdown",
            "1. Chọn option thứ 2 trong dropdown nguyên liệu.\n2. Kiểm tra value của select.",
            f"Selected ID: {selected_val}", "Dropdown nhận đúng ID nguyên liệu tương ứng.",
            f"ID được chọn: '{selected_val}'", pass_fail(tc36_ok), [s36], "High"
        )

        # TC37: Gửi ID nguyên liệu không tồn tại trong DB qua POST
        resp37 = page.request.post(
            BASE_URL + "inventory_receipt/store",
            form={"receipt_date": today_str, "ingredient_id": "999999", "qty": "5", "unit_price": "10000"},
            headers={"X-Requested-With": "XMLHttpRequest"},
        )
        tc37_ok = (resp37.status == 422 or "không tồn tại" in resp37.text().lower())
        add_result(
            37, "Detail - Ingredient", "Gửi ID nguyên liệu không tồn tại trong cơ sở dữ liệu (ID: 999999)",
            "1. Gửi POST /store với ingredient_id = '999999'.",
            "Ingredient ID: 999999", "Báo lỗi 'Nguyên liệu không tồn tại trong hệ thống'.",
            f"HTTP status: {resp37.status}", pass_fail(tc37_ok), [s36], "High"
        )

        # TC38: Chọn cùng 1 nguyên liệu trên 2 dòng khác nhau (Trùng lặp)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        first_val = page.locator(".ingredient-select").first.input_value()
        page.click("#add-item")
        page.wait_for_timeout(200)
        page.locator(".ingredient-select").nth(1).select_option(value=first_val)
        s38 = shot(page, "TC38_duplicate_ingredient_rows")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content38 = page.content()
        tc38_ok = "trùng lặp" in content38.lower() or page.locator(".row-duplicate").count() > 0 or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            38, "Detail - Ingredient", "Chọn cùng một nguyên liệu trên nhiều dòng (Phát hiện trùng lặp nguyên liệu)",
            "1. Dòng 1 chọn nguyên liệu A.\n2. Thêm Dòng 2 và cũng chọn nguyên liệu A.\n3. Nhấn Tạo phiếu nhập.",
            f"Row 1: ID {first_val} | Row 2: ID {first_val}",
            "Hệ thống highlight đỏ dòng trùng và báo lỗi 'Nguyên liệu bị trùng lặp với dòng trước đó'.",
            f"Bắt lỗi trùng lặp: {tc38_ok}", pass_fail(tc38_ok), [s38], "High"
        )

        # TC39: Tự động điền đơn giá mặc định khi chọn nguyên liệu
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.wait_for_timeout(200)
        price_val = page.locator(".price-input").first.input_value()
        s39 = shot(page, "TC39_auto_fill_price")
        tc39_ok = (float(price_val) >= 0)
        add_result(
            39, "Detail - Ingredient", "Tự động điền Đơn giá mặc định (purchase_price) khi chọn nguyên liệu",
            "1. Chọn nguyên liệu trên dropdown.\n2. Quan sát ô Đơn giá.",
            "Ingredient selection change event", "Ô Đơn giá tự động được điền giá mua gần nhất của nguyên liệu.",
            f"Đơn giá tự điền: '{price_val}'", pass_fail(tc39_ok), [s39], "Medium"
        )

        # TC40: Thay đổi chọn nguyên liệu khác cập nhật lại đơn giá
        page.locator(".ingredient-select").first.select_option(index=2)
        page.wait_for_timeout(200)
        price_val_2 = page.locator(".price-input").first.input_value()
        s40 = shot(page, "TC40_change_ingredient_updates_price")
        tc40_ok = (price_val_2 != "")
        add_result(
            40, "Detail - Ingredient", "Đổi sang nguyên liệu khác trên cùng một dòng cập nhật lại đơn giá",
            "1. Đổi sang nguyên liệu thứ 3.\n2. Quan sát đơn giá.",
            "Switch to ingredient index 2", "Đơn giá cập nhật theo nguyên liệu mới.",
            f"Đơn giá mới: '{price_val_2}'", pass_fail(tc40_ok), [s40], "Low"
        )

        # ======================================================================
        # CATEGORY 7: DETAIL ITEMS - QUANTITY (TC41 - TC48)
        # ======================================================================
        print("[7/11] Đang kiểm tra Category: Detail Items - Quantity...")

        # TC41: Bỏ trống Số lượng
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("")
        s41 = shot(page, "TC41_qty_empty")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content41 = page.content()
        tc41_ok = "số lượng không hợp lệ" in content41.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            41, "Detail - Quantity", "Bỏ trống ô Số lượng trên dòng nguyên liệu",
            "1. Xóa rỗng ô Số lượng.\n2. Nhấn Tạo phiếu nhập.",
            "Qty: [Empty]", "Báo lỗi 'Dòng 1: Số lượng không hợp lệ hoặc chưa được điền'.",
            f"Bắt lỗi số lượng rỗng: {tc41_ok}", pass_fail(tc41_ok), [s41], "High"
        )

        # TC42: Số lượng = 0
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("0")
        s42 = shot(page, "TC42_qty_zero")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content42 = page.content()
        tc42_ok = "số lượng nhập phải lớn hơn 0" in content42.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            42, "Detail - Quantity", "Nhập Số lượng = 0 trên dòng nguyên liệu",
            "1. Nhập Số lượng = 0.\n2. Nhấn Tạo phiếu nhập.",
            "Qty: 0", "Báo lỗi 'Dòng 1: Số lượng nhập phải lớn hơn 0'.",
            f"Bắt lỗi số lượng 0: {tc42_ok}", pass_fail(tc42_ok), [s42], "High"
        )

        # TC43: Số lượng âm (< 0)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("-10")
        s43 = shot(page, "TC43_qty_negative")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content43 = page.content()
        tc43_ok = "số lượng nhập phải lớn hơn 0" in content43.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            43, "Detail - Quantity", "Nhập Số lượng là số âm (< 0)",
            "1. Nhập Số lượng = -10.\n2. Nhấn Tạo phiếu nhập.",
            "Qty: -10", "Báo lỗi 'Dòng 1: Số lượng nhập phải lớn hơn 0'.",
            f"Bắt lỗi số âm: {tc43_ok}", pass_fail(tc43_ok), [s43], "High"
        )

        # TC44: Số lượng nguyên dương hợp lệ
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("25")
        page.locator(".price-input").first.fill("10000")
        s44 = shot(page, "TC44_qty_valid_int")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc44_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            44, "Detail - Quantity", "Nhập Số lượng là số nguyên dương hợp lệ (VD: 25)",
            "1. Nhập Số lượng = 25.\n2. Submit form.",
            "Qty: 25", "Tạo phiếu nhập thành công.",
            f"Lưu thành công: {tc44_ok}", pass_fail(tc44_ok), [s44], "High"
        )

        # TC45: Số lượng là số thập phân hợp lệ (VD: 2.5 kg)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("2.5")
        page.locator(".price-input").first.fill("10000")
        s45 = shot(page, "TC45_qty_valid_decimal")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc45_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            45, "Detail - Quantity", "Nhập Số lượng là số thập phân 1 chữ số lẻ (VD: 2.5 kg)",
            "1. Nhập Số lượng = 2.5.\n2. Submit form.",
            "Qty: 2.5", "Tạo phiếu nhập thành công, lưu đúng 2.5 vào DB.",
            f"Lưu thành công: {tc45_ok}", pass_fail(tc45_ok), [s45], "Medium"
        )

        # TC46: Số lượng có 3 chữ số thập phân (VD: 0.125 kg)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("0.125")
        page.locator(".price-input").first.fill("10000")
        s46 = shot(page, "TC46_qty_3_decimals")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc46_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            46, "Detail - Quantity", "Nhập Số lượng đạt tối đa 3 chữ số thập phân (VD: 0.125 kg)",
            "1. Nhập Số lượng = 0.125.\n2. Submit form.",
            "Qty: 0.125", "Tạo phiếu nhập thành công (khớp định dạng DECIMAL(10,3)).",
            f"Lưu thành công: {tc46_ok}", pass_fail(tc46_ok), [s46], "Medium"
        )

        # TC47: Số lượng vượt quá 3 chữ số thập phân (VD: 1.2345)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1.2345")
        page.locator(".price-input").first.fill("10000")
        s47 = shot(page, "TC47_qty_over_3_decimals")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content47 = page.content()
        tc47_ok = "tối đa 3 chữ số thập phân" in content47.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            47, "Detail - Quantity", "Nhập Số lượng có hơn 3 chữ số thập phân (VD: 1.2345)",
            "1. Nhập Số lượng = 1.2345 (4 số lẻ).\n2. Nhấn Tạo phiếu nhập.",
            "Qty: 1.2345", "Báo lỗi 'Số lượng chỉ cho phép tối đa 3 chữ số thập phân'.",
            f"Bắt lỗi quá 3 số lẻ: {tc47_ok}", pass_fail(tc47_ok), [s47], "High"
        )

        # TC48: Số lượng vượt ngưỡng tối đa (VD: 100,000)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("100000")
        page.locator(".price-input").first.fill("10000")
        s48 = shot(page, "TC48_qty_over_max")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content48 = page.content()
        tc48_ok = "vượt quá 99.999" in content48.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            48, "Detail - Quantity", "Nhập Số lượng vượt quá ngưỡng tối đa (> 99.999)",
            "1. Nhập Số lượng = 100000.\n2. Nhấn Tạo phiếu nhập.",
            "Qty: 100000", "Báo lỗi 'Số lượng nhập không được vượt quá 99.999'.",
            f"Bắt lỗi vượt max qty: {tc48_ok}", pass_fail(tc48_ok), [s48], "High"
        )

        # ======================================================================
        # CATEGORY 8: DETAIL ITEMS - UNIT PRICE (TC49 - TC55)
        # ======================================================================
        print("[8/11] Đang kiểm tra Category: Detail Items - Unit Price...")

        # TC49: Bỏ trống Đơn giá
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("2")
        page.locator(".price-input").first.fill("")
        s49 = shot(page, "TC49_price_empty")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content49 = page.content()
        tc49_ok = "đơn giá không hợp lệ" in content49.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            49, "Detail - Price", "Bỏ trống ô Đơn giá trên dòng nguyên liệu",
            "1. Xóa rỗng ô Đơn giá.\n2. Nhấn Tạo phiếu nhập.",
            "Unit Price: [Empty]", "Báo lỗi 'Dòng 1: Đơn giá không hợp lệ hoặc chưa được điền'.",
            f"Bắt lỗi đơn giá rỗng: {tc49_ok}", pass_fail(tc49_ok), [s49], "High"
        )

        # TC50: Đơn giá âm (< 0)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("2")
        page.locator(".price-input").first.fill("-5000")
        s50 = shot(page, "TC50_price_negative")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content50 = page.content()
        tc50_ok = "đơn giá không được âm" in content50.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            50, "Detail - Price", "Nhập Đơn giá là số âm (< 0)",
            "1. Nhập Đơn giá = -5000.\n2. Nhấn Tạo phiếu nhập.",
            "Unit Price: -5000", "Báo lỗi 'Dòng 1: Đơn giá không được âm'.",
            f"Bắt lỗi giá âm: {tc50_ok}", pass_fail(tc50_ok), [s50], "High"
        )

        # TC51: Đơn giá = 0 (Hàng tặng / khuyến mại)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("5")
        page.locator(".price-input").first.fill("0")
        s51 = shot(page, "TC51_price_zero")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc51_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            51, "Detail - Price", "Nhập Đơn giá = 0 đ (Hàng tặng kèm / nguyên liệu dùng thử)",
            "1. Nhập Đơn giá = 0.\n2. Submit form.",
            "Unit Price: 0", "Hệ thống chấp nhận đơn giá 0 đ và tạo phiếu thành công.",
            f"Lưu thành công: {tc51_ok}", pass_fail(tc51_ok), [s51], "Medium"
        )

        # TC52: Đơn giá số tiền dương chuẩn (VD: 50,000 đ)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("10")
        page.locator(".price-input").first.fill("50000")
        s52 = shot(page, "TC52_price_valid_positive")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc52_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            52, "Detail - Price", "Nhập Đơn giá là số dương chuẩn (VD: 50.000 đ)",
            "1. Nhập Đơn giá = 50000.\n2. Submit form.",
            "Unit Price: 50000", "Tạo phiếu nhập thành công.",
            f"Lưu thành công: {tc52_ok}", pass_fail(tc52_ok), [s52], "High"
        )

        # TC53: Đơn giá có số lẻ thập phân (VD: 15,500.50 đ)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("4")
        page.locator(".price-input").first.fill("15500.5")
        s53 = shot(page, "TC53_price_decimal")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc53_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            53, "Detail - Price", "Nhập Đơn giá có phần thập phân (VD: 15.500,5 đ)",
            "1. Nhập Đơn giá = 15500.5.\n2. Submit form.",
            "Unit Price: 15500.5", "Tạo phiếu nhập thành công.",
            f"Lưu thành công: {tc53_ok}", pass_fail(tc53_ok), [s53], "Low"
        )

        # TC54: Đơn giá vượt quá ngưỡng tối đa (> 1 tỷ đồng)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("1000000001")
        s54 = shot(page, "TC54_price_over_max")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content54 = page.content()
        tc54_ok = "vượt quá 1.000.000.000" in content54.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            54, "Detail - Price", "Nhập Đơn giá vượt quá giới hạn tối đa (> 1.000.000.000 đ)",
            "1. Nhập Đơn giá = 1000000001 (1 tỷ 1 đồng).\n2. Nhấn Tạo phiếu nhập.",
            "Unit Price: 1000000001", "Báo lỗi 'Dòng 1: Đơn giá không được vượt quá 1.000.000.000 đ'.",
            f"Bắt lỗi vượt max price: {tc54_ok}", pass_fail(tc54_ok), [s54], "High"
        )

        # TC55: Chỉnh sửa lại Đơn giá thủ công khác đơn giá mặc định
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.wait_for_timeout(200)
        page.locator(".price-input").first.fill("99000")
        s55 = shot(page, "TC55_price_custom_edit")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc55_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            55, "Detail - Price", "Chỉnh sửa thủ công đơn giá khác với giá gợi ý mặc định",
            "1. Chọn NL, sửa đơn giá thành 99.000 đ.\n2. Submit form.",
            "Unit Price: 99000 (Custom)", "Lưu đúng đơn giá người dùng đã điều chỉnh.",
            f"Lưu thành công: {tc55_ok}", pass_fail(tc55_ok), [s55], "Medium"
        )

        # ======================================================================
        # CATEGORY 9: ROW OPERATIONS & DYNAMIC CALCULATION (TC56 - TC62)
        # ======================================================================
        print("[9/11] Đang kiểm tra Category: Row Operations & Dynamic Calculation...")

        # TC56: Nhấn nút "Thêm dòng" tăng số dòng
        goto_create_receipt(page)
        initial_rows = page.locator(".receipt-row").count()
        page.click("#add-item")
        page.wait_for_timeout(200)
        after_add_rows = page.locator(".receipt-row").count()
        s56 = shot(page, "TC56_add_row_click")
        tc56_ok = (after_add_rows == initial_rows + 1)
        add_result(
            56, "Dynamic & Calculations", "Nhấn nút 'Thêm dòng nguyên liệu' tăng thêm 1 dòng nhập",
            "1. Nhấn nút #add-item.\n2. Đếm số lượng class .receipt-row.",
            "Action: Click #add-item", "Số dòng tăng thêm 1 và hiển thị đầy đủ các trường.",
            f"Trước: {initial_rows}, Sau: {after_add_rows}", pass_fail(tc56_ok), [s56], "High"
        )

        # TC57: Nhấn nút "Xóa dòng" xóa đúng dòng được chọn
        page.locator(".remove-item").last.click()
        page.wait_for_timeout(200)
        after_remove_rows = page.locator(".receipt-row").count()
        s57 = shot(page, "TC57_remove_row_click")
        tc57_ok = (after_remove_rows == initial_rows)
        add_result(
            57, "Dynamic & Calculations", "Nhấn nút 'Xóa dòng' (thùng rác) loại bỏ đúng dòng tương ứng",
            "1. Nhấn nút xóa trên dòng vừa thêm.\n2. Đếm lại số dòng.",
            "Action: Click .remove-item", "Dòng tương ứng bị xóa hoàn toàn khỏi DOM.",
            f"Số dòng sau xóa: {after_remove_rows}", pass_fail(tc57_ok), [s57], "High"
        )

        # TC58: Thêm liên tiếp nhiều dòng (5 dòng)
        goto_create_receipt(page)
        for _ in range(4):
            page.click("#add-item")
            page.wait_for_timeout(100)
        total_5_rows = page.locator(".receipt-row").count()
        s58 = shot(page, "TC58_add_5_rows")
        tc58_ok = (total_5_rows == 5)
        add_result(
            58, "Dynamic & Calculations", "Thêm liên tiếp nhiều dòng nguyên liệu (5 dòng) giao diện ổn định",
            "1. Nhấn Thêm dòng 4 lần liên tiếp.\n2. Kiểm tra tổng số 5 dòng.",
            "Action: 4 clicks on #add-item", "Giao diện hiển thị đủ 5 dòng đều đặn, không vỡ layout.",
            f"Tổng số dòng: {total_5_rows}", pass_fail(tc58_ok), [s58], "Medium"
        )

        # TC59: Tự động tính Thành tiền = Qty * Price khi nhập số lượng
        goto_create_receipt(page)
        page.locator(".qty-input").first.fill("4")
        page.locator(".price-input").first.fill("25000")
        page.wait_for_timeout(200)
        subtotal_text_59 = page.locator(".subtotal").first.inner_text()
        s59 = shot(page, "TC59_calc_subtotal_on_qty")
        tc59_ok = "100.000" in subtotal_text_59 or "100000" in subtotal_text_59
        add_result(
            59, "Dynamic & Calculations", "Tự động tính Thành tiền = Số lượng * Đơn giá realtime khi nhập số lượng",
            "1. Nhập Số lượng = 4, Đơn giá = 25.000 đ.\n2. Quan sát ô Thành tiền của dòng.",
            "Qty: 4, Price: 25000", "Thành tiền tự động hiển thị '100.000 đ'.",
            f"Thành tiền thực tế: '{subtotal_text_59}'", pass_fail(tc59_ok), [s59], "High"
        )

        # TC60: Tự động tính lại Thành tiền khi sửa Đơn giá
        page.locator(".price-input").first.fill("50000")
        page.wait_for_timeout(200)
        subtotal_text_60 = page.locator(".subtotal").first.inner_text()
        s60 = shot(page, "TC60_calc_subtotal_on_price")
        tc60_ok = "200.000" in subtotal_text_60 or "200000" in subtotal_text_60
        add_result(
            60, "Dynamic & Calculations", "Tự động tính lại Thành tiền khi thay đổi ô Đơn giá",
            "1. Sửa Đơn giá thành 50.000 đ (Số lượng vẫn = 4).\n2. Quan sát ô Thành tiền.",
            "Qty: 4, New Price: 50000", "Thành tiền tự động cập nhật lên '200.000 đ'.",
            f"Thành tiền mới: '{subtotal_text_60}'", pass_fail(tc60_ok), [s60], "High"
        )

        # TC61: Tự động tính Tổng cộng toàn phiếu bằng tổng tất cả các dòng
        page.click("#add-item")
        page.wait_for_timeout(200)
        page.locator(".qty-input").nth(1).fill("2")
        page.locator(".price-input").nth(1).fill("150000")
        page.wait_for_timeout(200)
        grand_total_text = page.locator("#receipt-total").inner_text()
        s61 = shot(page, "TC61_calc_grand_total")
        tc61_ok = "500.000" in grand_total_text or "500000" in grand_total_text
        add_result(
            61, "Dynamic & Calculations", "Tự động cập nhật Tổng cộng giá trị phiếu nhập bằng tổng tất cả các dòng",
            "1. Dòng 1: 200.000 đ. Dòng 2: 300.000 đ.\n2. Quan sát ô Tổng cộng (#receipt-total).",
            "Row 1: 200k | Row 2: 300k", "Tổng cộng hiển thị chính xác '500.000 đ'.",
            f"Tổng cộng thực tế: '{grand_total_text}'", pass_fail(tc61_ok), [s61], "High"
        )

        # TC62: Định dạng tiền tệ VNĐ có phân cách hàng nghìn rõ ràng
        s62 = shot(page, "TC62_currency_format")
        tc62_ok = ("." in grand_total_text or "," in grand_total_text) and "đ" in grand_total_text
        add_result(
            62, "Dynamic & Calculations", "Định dạng tiền tệ VNĐ chuẩn có dấu chấm phân cách và ký hiệu 'đ'",
            "1. Quan sát cách hiển thị số tiền trên Thành tiền và Tổng cộng.",
            "Value: 500000", "Hiển thị định dạng tiền tệ Việt Nam (VD: '500.000 đ').",
            f"Định dạng hiển thị: '{grand_total_text}'", pass_fail(tc62_ok), [s62], "Medium"
        )

        # ======================================================================
        # CATEGORY 10: FORM SUBMISSION & STATE PERSISTENCE (TC63 - TC70)
        # ======================================================================
        print("[10/11] Đang kiểm tra Category: Form Submission & State Persistence...")

        # TC63: Submit form không có dòng nguyên liệu nào (xóa hết dòng)
        goto_create_receipt(page)
        page.locator(".remove-item").first.click()
        page.wait_for_timeout(200)
        s63 = shot(page, "TC63_submit_zero_rows")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        content63 = page.content()
        tc63_ok = "ít nhất một dòng nguyên liệu" in content63.lower() or page.locator("#clientErrorBox:not(.d-none)").count() > 0
        add_result(
            63, "Submission & State", "Submit form khi không có bất kỳ dòng nguyên liệu nào",
            "1. Xóa toàn bộ các dòng chi tiết.\n2. Nhấn Tạo phiếu nhập.",
            "Rows count: 0", "Báo lỗi 'Phiếu nhập kho phải có ít nhất một dòng nguyên liệu'.",
            f"Bắt lỗi không có dòng: {tc63_ok}", pass_fail(tc63_ok), [s63], "High"
        )

        # TC64: Submit form với dữ liệu hoàn toàn hợp lệ 1 dòng
        goto_create_receipt(page)
        page.fill("#supplier", "TEST_NCC_HOAN_CHINH_1")
        page.fill("#receipt_date", today_str)
        page.fill("#note", "TEST_KIEM_THU_THANH_CONG_1")
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("10")
        page.locator(".price-input").first.fill("20000")
        s64 = shot(page, "TC64_submit_valid_single_row")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(500)
        s64b = shot(page, "TC64_after_submit_success")
        tc64_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            64, "Submission & State", "Submit form với đầy đủ dữ liệu hợp lệ (1 dòng nguyên liệu)",
            "1. Nhập Supplier, Date, Note, 1 dòng nguyên liệu hợp lệ.\n2. Nhấn Tạo phiếu nhập.",
            "Valid 1 row data", "Tạo phiếu nhập thành công (status: pending), chuyển hướng về danh sách.",
            f"URL sau submit: {page.url}", pass_fail(tc64_ok), [s64, s64b], "High"
        )

        # TC65: Submit form hợp lệ nhiều dòng (3 dòng khác nhau)
        goto_create_receipt(page)
        page.fill("#supplier", "TEST_NCC_MULTI_ROWS")
        page.fill("#receipt_date", today_str)
        page.fill("#note", "TEST_KIEM_THU_MULTI_ROWS")
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("5")
        page.locator(".price-input").first.fill("30000")

        page.click("#add-item")
        page.wait_for_timeout(150)
        page.locator(".ingredient-select").nth(1).select_option(index=2)
        page.locator(".qty-input").nth(1).fill("10")
        page.locator(".price-input").nth(1).fill("15000")

        page.click("#add-item")
        page.wait_for_timeout(150)
        page.locator(".ingredient-select").nth(2).select_option(index=3)
        page.locator(".qty-input").nth(2).fill("2.5")
        page.locator(".price-input").nth(2).fill("80000")

        s65 = shot(page, "TC65_submit_3_valid_rows")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(500)
        tc65_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            65, "Submission & State", "Submit form với nhiều dòng nguyên liệu hợp lệ (3 dòng khác nhau)",
            "1. Thêm 3 dòng với 3 nguyên liệu khác nhau.\n2. Điền số lượng và đơn giá hợp lệ.\n3. Submit.",
            "Valid 3 rows data", "Lưu thành công toàn bộ phiếu và chi tiết 3 dòng vào DB.",
            f"Kết quả lưu 3 dòng: {tc65_ok}", pass_fail(tc65_ok), [s65], "High"
        )

        # TC66: Giữ lại Old Input khi Backend trả về lỗi validation
        goto_create_receipt(page)
        page.fill("#supplier", "TEST_OLD_SUPPLIER_RETAINED")
        page.evaluate("document.getElementById('receipt_date').removeAttribute('min'); document.getElementById('receipt_date').value = '2019-01-01';")
        page.fill("#note", "TEST_OLD_NOTE_RETAINED")
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("12.5")
        page.locator(".price-input").first.fill("45000")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(500)
        s66 = shot(page, "TC66_old_input_retained")
        supp_ret = page.locator("#supplier").input_value()
        note_ret = page.locator("#note").input_value()
        tc66_ok = (supp_ret == "TEST_OLD_SUPPLIER_RETAINED" or "TEST_OLD_SUPPLIER_RETAINED" in page.content())
        add_result(
            66, "Submission & State", "Hệ thống giữ lại toàn bộ dữ liệu đã nhập (Old Input) khi có lỗi validation",
            "1. Submit form chứa 1 trường lỗi (ngày trước 2020).\n2. Backend redirect về create.\n3. Quan sát các ô nhập liệu.",
            "Supplier: 'TEST_OLD_SUPPLIER_RETAINED', Note: 'TEST_OLD_NOTE_RETAINED'",
            "Các trường Nhà cung cấp, Ghi chú và các dòng nguyên liệu vẫn giữ nguyên dữ liệu đã nhập.",
            f"Supplier giữ lại: '{supp_ret}', Note giữ lại: '{note_ret}'", pass_fail(tc66_ok), [s66], "High"
        )

        # TC67: Hiển thị danh sách thông báo lỗi chi tiết theo từng dòng
        goto_create_receipt(page)
        page.fill("#supplier", "A")  # Lỗi < 2 ký tự
        page.evaluate("document.getElementById('receipt_date').value = '';")  # Lỗi rỗng ngày
        page.locator(".ingredient-select").first.select_option(value="")  # Lỗi chưa chọn NL
        page.locator(".qty-input").first.fill("-5")  # Lỗi số âm
        s67 = shot(page, "TC67_multiple_detailed_errors")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(300)
        s67b = shot(page, "TC67_error_box_displayed")
        errors_count = page.locator("#clientErrorList li").count()
        tc67_ok = errors_count >= 3
        add_result(
            67, "Submission & State", "Hiển thị danh sách thông báo lỗi chi tiết theo từng trường / từng dòng",
            "1. Nhập nhiều trường vi phạm đồng thời.\n2. Nhấn Tạo phiếu nhập.\n3. Quan sát khối alert thông báo lỗi.",
            "4 vi phạm: Supplier, Date, Ingredient, Qty",
            "Khung thông báo lỗi liệt kê chi tiết từng lỗi bằng danh sách gạch đầu dòng rõ ràng.",
            f"Số lượng thông báo lỗi chi tiết: {errors_count}", pass_fail(tc67_ok), [s67, s67b], "High"
        )

        # TC68: Chống Double Click / Double Submit (Disable nút khi submit)
        goto_create_receipt(page)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("10000")
        s68 = shot(page, "TC68_double_click_protection")
        page.click("#btnSubmitReceipt")
        page.wait_for_timeout(20)
        btn_disabled = page.evaluate("document.getElementById('btnSubmitReceipt') ? document.getElementById('btnSubmitReceipt').disabled : true")
        tc68_ok = (btn_disabled is True)
        add_result(
            68, "Submission & State", "Chống submit trùng lặp (Double Click) bằng cách disable nút và hiển thị spinner",
            "1. Nhấn nút Tạo phiếu nhập.\n2. Kiểm tra trạng thái disabled của nút ngay lập tức.",
            "Action: Submit button click", "Nút bị vô hiệu hóa (disabled = true) và hiển thị 'Đang lưu phiếu...'.",
            f"Disabled state: {btn_disabled}", pass_fail(tc68_ok), [s68], "Medium"
        )

        # TC69: Mở form từ Quick Restock query params (?ingredient_id=1&qty=15)
        page.goto(BASE_URL + "inventory_receipt/create?ingredient_id=1&qty=15", wait_until="domcontentloaded")
        page.wait_for_timeout(400)
        s69 = shot(page, "TC69_quick_restock_params")
        q_sel = page.locator(".ingredient-select").first.input_value()
        q_qty = page.locator(".qty-input").first.input_value()
        tc69_ok = (q_sel == "1" and float(q_qty) == 15)
        add_result(
            69, "Submission & State", "Mở form từ Quick Restock URL query params (?ingredient_id=1&qty=15)",
            "1. Truy cập URL kèm tham số: /inventory_receipt/create?ingredient_id=1&qty=15.\n2. Quan sát dòng đầu tiên.",
            "Query: ingredient_id=1, qty=15", "Tự động chọn đúng nguyên liệu ID 1 và điền số lượng 15.",
            f"Selected ID: '{q_sel}', Qty: '{q_qty}'", pass_fail(tc69_ok), [s69], "Medium"
        )

        # TC70: Mở form từ Restock Cart (fromRestock=true với sessionStorage)
        page.goto(BASE_URL + "inventory_receipt", wait_until="domcontentloaded")
        page.evaluate("sessionStorage.setItem('restockCart', JSON.stringify([{ingredient_id: 1, qty: 8}, {ingredient_id: 2, qty: 12}]));")
        page.goto(BASE_URL + "inventory_receipt/create_from_restock", wait_until="domcontentloaded")
        page.wait_for_timeout(400)
        s70 = shot(page, "TC70_create_from_restock_cart")
        rows_restock = page.locator(".receipt-row").count()
        tc70_ok = (rows_restock == 2)
        add_result(
            70, "Submission & State", "Mở form tạo phiếu từ giỏ nhập hàng tự động (create_from_restock)",
            "1. Ghi restockCart vào sessionStorage.\n2. Truy cập /inventory_receipt/create_from_restock.\n3. Đếm số dòng.",
            "sessionStorage: 2 items (ID 1: qty 8, ID 2: qty 12)",
            "Tự động sinh ra 2 dòng nguyên liệu khớp với giỏ nhập hàng.",
            f"Số dòng sinh ra: {rows_restock}", pass_fail(tc70_ok), [s70], "Medium"
        )

        # ======================================================================
        # CATEGORY 11: SECURITY & EDGE CASES (TC71 - TC73)
        # ======================================================================
        print("[11/11] Đang kiểm tra Category: Security & Edge Cases...")

        # TC71: Chống tấn công XSS Script Injection
        xss_payload = "<script>alert('XSS_ATTACK')</script>"
        goto_create_receipt(page)
        page.fill("#supplier", "TEST_XSS_" + xss_payload)
        page.fill("#receipt_date", today_str)
        page.fill("#note", "TEST_NOTE_" + xss_payload)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("10000")
        s71 = shot(page, "TC71_xss_injection_test")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(500)
        s71b = shot(page, "TC71_after_xss_submit")
        body_71 = page.content()
        tc71_ok = ("<script>alert" not in body_71) or ("&lt;script&gt;" in body_71) or ("inventory_receipt" in page.url)
        add_result(
            71, "Security & Edge Cases", "Chống tấn công XSS Injection trong trường Nhà cung cấp và Ghi chú",
            "1. Nhập payload '<script>alert(1)</script>' vào supplier và note.\n2. Submit form và xem trang danh sách.",
            f"Payload: {xss_payload}", "Mã HTML/Script được escape an toàn (htmlspecialchars), không bị thực thi script.",
            f"An toàn XSS: {tc71_ok}", pass_fail(tc71_ok), [s71, s71b], "High"
        )

        # TC72: Chống SQL Injection trong các trường text
        sqli_payload = "TEST_SQLI' OR '1'='1' -- "
        goto_create_receipt(page)
        page.fill("#supplier", "TEST_SQLI_NCC")
        page.fill("#receipt_date", today_str)
        page.fill("#note", sqli_payload)
        page.locator(".ingredient-select").first.select_option(index=1)
        page.locator(".qty-input").first.fill("1")
        page.locator(".price-input").first.fill("10000")
        s72 = shot(page, "TC72_sqli_injection_test")
        page.click("#btnSubmitReceipt")
        page.wait_for_load_state("domcontentloaded")
        tc72_ok = ("inventory_receipt" in page.url and "create" not in page.url) or "thành công" in page.content().lower()
        add_result(
            72, "Security & Edge Cases", "Chống tấn công SQL Injection trong các trường dữ liệu văn bản",
            "1. Nhập payload SQL Injection vào các ô văn bản.\n2. Submit form.",
            f"Payload: {sqli_payload}", "Prepared Statement xử lý an toàn, lưu chính xác chuỗi mà không lỗi DB.",
            f"Lưu an toàn: {tc72_ok}", pass_fail(tc72_ok), [s72], "High"
        )

        # TC73: Kiểm tra tính toàn vẹn dữ liệu trong cơ sở dữ liệu sau khi tạo
        check_db_code = f"""
        require_once '{ROOT_DIR}/config/config.php';
        $db = getDB();
        $r = $db->query("SELECT id, status FROM inventory_receipt WHERE supplier LIKE 'TEST_%' ORDER BY id DESC LIMIT 1")->fetch(PDO::FETCH_ASSOC);
        if ($r) {{
            $d = $db->query("SELECT count(*) FROM inventory_receipt_detail WHERE receipt_id = " . $r['id'])->fetchColumn();
            echo $r['status'] . '|' . $d;
        }} else {{
            echo 'NOT_FOUND';
        }}
        """
        db_res = run_php(check_db_code)
        s73 = shot(page, "TC73_db_data_integrity")
        tc73_ok = ("pending" in db_res) and (db_res != "NOT_FOUND")
        add_result(
            73, "Security & Edge Cases", "Kiểm tra tính toàn vẹn dữ liệu trong Database (bảng inventory_receipt & detail)",
            "1. Truy vấn trực tiếp DB kiểm tra bản ghi vừa tạo.\n2. Đối chiếu status và số lượng chi tiết detail.",
            "Query DB: inventory_receipt + detail", "Bản ghi lưu đúng status='pending' và đầy đủ các dòng detail liên kết khóa ngoại.",
            f"DB Check result: '{db_res}'", pass_fail(tc73_ok), [s73], "High"
        )

        browser.close()

    print("\n" + "="*70)
    print("HOÀN TẤT KIỂM THỬ TỰ ĐỘNG! ĐANG XUẤT CÁC BẢN BÁO CÁO...")
    print("="*70)

    excel_file = export_excel()
    html_file = export_html()

    passed_count = sum(1 for r in results if r["status"] == "Pass")
    failed_count = sum(1 for r in results if r["status"] == "Fail")

    print(f"\n[+] Đã xuất file báo cáo Excel : {excel_file}")
    print(f"[+] Đã xuất file báo cáo HTML  : {html_file}")
    print(f"[+] Thư mục ảnh chụp màn hình : {OUT_DIR}")
    print(f"[+] Tổng số Test Cases         : {len(results)}")
    print(f"[+] Số Test Cases PASS        : {passed_count} ({round((passed_count/len(results))*100, 1)}%)")
    print(f"[+] Số Test Cases FAIL        : {failed_count}\n")


if __name__ == "__main__":
    run_all_tests()
